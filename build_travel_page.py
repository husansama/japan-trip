from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("travel-page")
WORKBOOK = Path("日本12天11晚_0925-1006_P人关西加强版_含美食.xlsx")


def segment(
    start: str,
    end: str,
    duration: str,
    mode: str,
    *,
    option: str = "",
    note: str = "",
    map_start: str = "",
    map_end: str = "",
) -> dict[str, str]:
    """统一保存分段交通，并保留用于生成 Google Maps 实时路线的精确地点名。"""
    return {
        "起点": start,
        "终点": end,
        "耗时": duration,
        "方式": mode,
        "方案": option,
        "说明": note,
        "地图起点": map_start or start,
        "地图终点": map_end or end,
    }


# 耗时于 2026-07-30 按 Google Maps 日间推荐路线核对；班次、候车与拥堵会让结果浮动。
SEGMENT_ROUTES: dict[str, list[dict[str, str]]] = {
    "Day1": [
        segment("关西机场 KIX", "大阪站/梅田", "约47分钟", "公共交通", option="住梅田", note="HARUKA 直达", map_start="Kansai International Airport", map_end="Osaka Station"),
        segment("关西机场 KIX", "难波站", "约46分钟", "公共交通", option="住难波", note="南海机场线；特急班次可约42分钟", map_start="Kansai International Airport", map_end="Namba Station Osaka"),
        segment("大阪站/梅田", "道顿堀", "约19分钟", "公共交通", option="住梅田", note="御堂筋线到难波后步行", map_start="Osaka Station", map_end="Dotonbori Osaka"),
        segment("难波站", "道顿堀", "约5分钟", "步行", option="住难波", map_start="Namba Station Osaka", map_end="Dotonbori Osaka"),
    ],
    "Day2": [
        segment("大阪站", "大阪城", "约29分钟", "公共交通", note="含车站至天守阁区域步行", map_start="Osaka Station", map_end="Osaka Castle"),
        segment("大阪城", "梅田", "约28分钟", "公共交通", note="步行接谷町线", map_start="Osaka Castle", map_end="Umeda Station"),
        segment("梅田", "中崎町", "约12分钟", "步行", map_start="Umeda Station", map_end="Nakazakicho Osaka"),
        segment("中崎町", "心斋桥", "约19分钟", "公共交通", note="步行至梅田后乘御堂筋线", map_start="Nakazakicho Station", map_end="Shinsaibashi Station"),
        segment("心斋桥", "道顿堀", "约12分钟", "步行", map_start="Shinsaibashi Station", map_end="Dotonbori Osaka"),
        segment("道顿堀", "大阪站", "约18分钟", "公共交通", note="难波乘御堂筋线", map_start="Dotonbori Osaka", map_end="Osaka Station"),
    ],
    "Day3": [
        segment("大阪站", "姬路站", "约1小时2分钟", "公共交通", note="JR新快速，无需预约", map_start="Osaka Station", map_end="Himeji Station"),
        segment("姬路站", "姬路城", "约20分钟", "步行", map_start="Himeji Station", map_end="Himeji Castle"),
        segment("姬路城", "神户三宫", "约1小时1分钟", "公共交通", note="含步行回姬路站", map_start="Himeji Castle", map_end="Sannomiya Station Kobe"),
        segment("三宫站", "北野异人馆", "约16分钟", "步行", note="上坡路段", map_start="Sannomiya Station Kobe", map_end="Kitano Ijinkan Kobe"),
        segment("北野异人馆", "南京町", "约24分钟", "公共交通", note="公交加步行", map_start="Kitano Ijinkan Kobe", map_end="Nankinmachi Kobe"),
        segment("南京町", "Harborland", "约19分钟", "步行", map_start="Nankinmachi Kobe", map_end="Kobe Harborland"),
        segment("Harborland", "大阪站", "约34分钟", "公共交通", note="从神户站乘JR新快速", map_start="Kobe Harborland", map_end="Osaka Station"),
    ],
    "Day4": [
        segment("大阪站", "环球城站", "约11分钟", "公共交通", note="大阪环状线接JR梦咲线", map_start="Osaka Station", map_end="Universal City Station Osaka"),
        segment("环球城站", "大阪站", "约15分钟", "公共交通", note="散场高峰另预留进站时间", map_start="Universal City Station Osaka", map_end="Osaka Station"),
    ],
    "Day5": [
        segment("大阪站", "四条站", "约38分钟", "公共交通", note="当前最快路线经京都站换乌丸线；拖箱建议多留10分钟", map_start="Osaka Station", map_end="Shijo Station Kyoto"),
        segment("四条站", "清水寺", "约29分钟", "公共交通", note="公交加步行；拥堵时可能更久", map_start="Shijo Station Kyoto", map_end="Kiyomizu-dera Kyoto"),
        segment("清水寺", "三年坂", "约3分钟", "步行", map_start="Kiyomizu-dera Kyoto", map_end="Sannenzaka Kyoto"),
        segment("三年坂", "八坂神社", "约14分钟", "步行", map_start="Sannenzaka Kyoto", map_end="Yasaka Shrine Kyoto"),
        segment("八坂神社", "祇园白川", "约9分钟", "步行", map_start="Yasaka Shrine Kyoto", map_end="Gion Shirakawa Kyoto"),
        segment("祇园白川", "先斗町", "约6分钟", "步行", map_start="Gion Shirakawa Kyoto", map_end="Pontocho Alley Kyoto"),
        segment("先斗町", "四条站", "约15分钟", "步行", map_start="Pontocho Alley Kyoto", map_end="Shijo Station Kyoto"),
    ],
    "Day6": [
        segment("四条站", "阪急岚山站", "约18分钟", "公共交通", note="乌丸站乘阪急，桂站换乘", map_start="Shijo Station Kyoto", map_end="Arashiyama Station Hankyu"),
        segment("阪急岚山站", "天龙寺", "约13分钟", "步行", map_start="Arashiyama Station Hankyu", map_end="Tenryu-ji Kyoto"),
        segment("天龙寺", "竹林小径", "约1–5分钟", "步行", note="入口紧邻天龙寺北门", map_start="Tenryu-ji Kyoto", map_end="Arashiyama Bamboo Forest"),
        segment("竹林小径", "渡月桥", "约13分钟", "步行", map_start="Arashiyama Bamboo Forest", map_end="Togetsukyo Bridge"),
        segment("渡月桥", "金阁寺", "约50分钟", "公共交通", note="公交换乘，候车影响较大", map_start="Togetsukyo Bridge", map_end="Kinkaku-ji Kyoto"),
        segment("金阁寺", "龙安寺", "约11分钟", "公共交通", note="59路公交加步行", map_start="Kinkaku-ji Kyoto", map_end="Ryoan-ji Kyoto"),
        segment("龙安寺", "四条站", "约40分钟", "公共交通", note="公交接乌丸线", map_start="Ryoan-ji Kyoto", map_end="Shijo Station Kyoto"),
    ],
    "Day7": [
        segment("四条站", "JR宇治站", "约34分钟", "公共交通", note="乌丸线至京都站换JR奈良线", map_start="Shijo Station Kyoto", map_end="Uji Station JR"),
        segment("JR宇治站", "平等院", "约9分钟", "步行", map_start="Uji Station JR", map_end="Byodoin Kyoto"),
        segment("平等院", "JR宇治站", "约9分钟", "步行", map_start="Byodoin Kyoto", map_end="Uji Station JR"),
        segment("JR宇治站", "JR奈良站", "约44分钟", "公共交通", note="JR奈良线", map_start="Uji Station JR", map_end="Nara Station JR"),
        segment("JR奈良站", "奈良公园", "约22分钟", "步行", map_start="Nara Station JR", map_end="Nara Park"),
        segment("奈良公园巴士总站", "若草山北口", "约39分钟", "步行", note="公园内部范围很大，此段按明确入口核对", map_start="Nara Park Bus Terminal", map_end="若草山登山口 北ゲート"),
        segment("若草山一带", "四条站", "约1小时21分钟", "公共交通", note="步行至公交站后转近铁/地铁", map_start="Mount Wakakusa", map_end="Shijo Station Kyoto"),
    ],
    "Day8": [
        segment("四条站", "贵船口站", "约58分钟", "公共交通", option="方案A 贵船鞍马", note="公交接叡山电铁", map_start="Shijo Station Kyoto", map_end="Kibuneguchi Station"),
        segment("贵船口站", "贵船神社", "约8分钟", "公共交通", option="方案A 贵船鞍马", note="33路公交；步行约30分钟", map_start="Kibuneguchi Station", map_end="Kifune Shrine Kyoto"),
        segment("贵船神社", "鞍马站", "约21分钟", "公共交通", option="方案A 贵船鞍马", note="公交回贵船口后乘叡山电铁；沿公路步行约43分钟", map_start="Kifune Shrine Kyoto", map_end="Kurama Station Kyoto"),
        segment("鞍马站", "四条站", "约1小时4分钟", "公共交通", option="方案A 贵船鞍马", note="叡山电铁接公交/地铁", map_start="Kurama Station Kyoto", map_end="Shijo Station Kyoto"),
        segment("四条站", "三千院", "约1小时19分钟", "公共交通", option="方案B 大原", note="直达公交加步行", map_start="Shijo Station Kyoto", map_end="Sanzen-in Kyoto"),
        segment("三千院", "宝泉院", "约3分钟", "步行", option="方案B 大原", map_start="Sanzen-in Kyoto", map_end="Hosen-in Kyoto"),
        segment("宝泉院", "四条站", "约1小时5分钟", "公共交通", option="方案B 大原", note="步行至大原站乘17路", map_start="Hosen-in Kyoto", map_end="Shijo Station Kyoto"),
        segment("四条站", "伏见稻荷", "约17分钟", "公共交通", option="方案C 市区", note="乌丸线至京都站换JR奈良线", map_start="Shijo Station Kyoto", map_end="Fushimi Inari Taisha"),
        segment("伏见稻荷", "锦市场", "约24分钟", "公共交通", option="方案C 市区", note="京阪线加步行", map_start="Fushimi Inari Taisha", map_end="Nishiki Market Kyoto"),
        segment("锦市场", "新京极", "约2分钟", "步行", option="方案C 市区", map_start="Nishiki Market Kyoto", map_end="Shinkyogoku Shopping Street"),
        segment("新京极", "四条站", "约14分钟", "步行", option="方案C 市区", map_start="Shinkyogoku Shopping Street", map_end="Shijo Station Kyoto"),
    ],
    "Day9": [
        segment("四条站", "京都站", "约4分钟", "公共交通", note="乌丸线；另留进站与找站台时间", map_start="Shijo Station Kyoto", map_end="Kyoto Station"),
        segment("京都站", "东京站", "约2小时11分钟", "公共交通", note="东海道新干线当前最快班次", map_start="Kyoto Station", map_end="Tokyo Station"),
        segment("东京站", "新宿站", "约14分钟", "公共交通", note="JR中央线", map_start="Tokyo Station", map_end="Shinjuku Station"),
        segment("新宿站", "六本木Hills", "约18分钟", "公共交通", option="夜景A", note="大江户线加步行", map_start="Shinjuku Station", map_end="Roppongi Hills"),
        segment("六本木Hills", "新宿站", "约17分钟", "公共交通", option="夜景A", map_start="Roppongi Hills", map_end="Shinjuku Station"),
        segment("新宿站", "涩谷十字路口", "约8分钟", "公共交通", option="夜景B", note="JR湘南新宿线；山手线也可", map_start="Shinjuku Station", map_end="Shibuya Scramble Crossing"),
    ],
    "Day10": [
        segment("新宿站", "明治神宫", "约7分钟", "公共交通", note="到参宫桥一带；进入主殿另需步行", map_start="Shinjuku Station", map_end="Meiji Jingu"),
        segment("明治神宫主殿", "原宿站", "约10分钟", "步行", map_start="Meiji Jingu Main Hall", map_end="Harajuku Station"),
        segment("原宿站", "表参道站", "约13分钟", "步行", map_start="Harajuku Station", map_end="Omotesando Station"),
        segment("表参道站", "涩谷十字路口", "约17分钟", "步行", map_start="Omotesando Station", map_end="Shibuya Scramble Crossing"),
        segment("涩谷站", "新宿站", "约7分钟", "公共交通", note="JR山手线", map_start="Shibuya Station", map_end="Shinjuku Station"),
    ],
    "Day11": [
        segment("新宿站", "河口湖站", "约1小时46分钟", "公共交通", option="方案A 河口湖", note="Busta新宿高速巴士；需预订", map_start="Shinjuku Station", map_end="Kawaguchiko Station"),
        segment("河口湖站", "新宿站", "约1小时46分钟", "公共交通", option="方案A 河口湖", note="高速巴士；返程也建议预订", map_start="Kawaguchiko Station", map_end="Shinjuku Station"),
        segment("新宿站", "镰仓站", "约1小时", "公共交通", option="方案B 镰仓江之岛", note="JR湘南新宿线", map_start="Shinjuku Station", map_end="Kamakura Station"),
        segment("镰仓站", "江之岛站", "约25分钟", "公共交通", option="方案B 镰仓江之岛", note="江之电", map_start="Kamakura Station", map_end="Enoshima Station"),
        segment("江之岛站", "新宿站", "约1小时19分钟", "公共交通", option="方案B 镰仓江之岛", note="步行至片濑江之岛后乘小田急", map_start="Enoshima Station", map_end="Shinjuku Station"),
        segment("新宿站", "银座站", "约15分钟", "公共交通", option="方案C 东京市区", note="丸之内线", map_start="Shinjuku Station", map_end="Ginza Station"),
        segment("银座站", "浅草寺", "约20分钟", "公共交通", option="方案C 东京市区", note="银座线", map_start="Ginza Station", map_end="Senso-ji"),
        segment("浅草寺", "秋叶原站", "约11分钟", "公共交通", option="方案C 东京市区", note="步行至筑波快线浅草站", map_start="Senso-ji", map_end="Akihabara Station"),
        segment("秋叶原站", "新宿站", "约23分钟", "公共交通", option="方案C 东京市区", note="JR中央·总武线", map_start="Akihabara Station", map_end="Shinjuku Station"),
    ],
    "Day12": [
        segment("新宿站", "浅草寺", "约30分钟", "公共交通", option="上午选浅草", note="中央线接银座线", map_start="Shinjuku Station", map_end="Senso-ji"),
        segment("浅草寺", "成田机场", "约1小时12分钟", "公共交通", option="浅草→成田", note="浅草线/京成线；另留值机时间", map_start="Senso-ji", map_end="Narita International Airport"),
        segment("浅草寺", "羽田T3", "约40分钟", "公共交通", option="浅草→羽田", note="浅草线直通；另留值机时间", map_start="Senso-ji", map_end="Haneda Airport Terminal 3"),
        segment("新宿站", "银座站", "约15分钟", "公共交通", option="上午选银座", note="丸之内线", map_start="Shinjuku Station", map_end="Ginza Station"),
        segment("银座站", "成田机场", "约1小时13分钟", "公共交通", option="银座→成田", note="当前推荐为机场巴士；另留值机时间", map_start="Ginza Station", map_end="Narita International Airport"),
        segment("银座站", "羽田T3", "约25分钟", "公共交通", option="银座→羽田", note="有乐町换山手线及东京单轨；另留值机时间", map_start="Ginza Station", map_end="Haneda Airport Terminal 3"),
    ],
}


def food_pin(name: str, query: str, note: str = "") -> dict[str, str]:
    """保存已核对的店名和地图检索词，前端据此生成可直接点击的 Google Maps 定位。"""
    return {"店名": name, "地图查询": query, "定位说明": note}


# 原攻略中部分条目是区域型推荐；这里为每类补充可落地导航的代表店或明确地点。
FOOD_LOCATIONS: dict[tuple[str, str], list[dict[str, str]]] = {
    ("大阪难波/道顿堀", "大阪烧"): [
        food_pin("美津の", "Mizuno Okonomiyaki Dotonbori Osaka"),
        food_pin("味乃家本店", "Ajinoya Honten Osaka"),
        food_pin("千房道顿堀", "Chibo Dotonbori Building"),
    ],
    ("大阪难波/道顿堀", "章鱼烧"): [
        food_pin("道顿堀くくる", "Dotonbori Kukuru Konamon Museum"),
        food_pin("十八番道顿堀店", "Takoyaki Juhachiban Dotonbori"),
    ],
    ("大阪新世界/难波", "串炸"): [
        food_pin("串かつだるま新世界总本店", "Kushikatsu Daruma Shinsekai Sohonten"),
    ],
    ("大阪梅田", "百货美食"): [
        food_pin("阪神百货梅田本店", "Hanshin Department Store Umeda Main Store"),
        food_pin("KITTE Osaka", "KITTE Osaka"),
    ],
    ("大阪梅田/天满", "寿司/居酒屋"): [
        food_pin("春驹本店", "Harukoma Honten Osaka"),
        food_pin("寿司酒场さしす梅田", "Sushi Sakaba Sashisu Umeda"),
    ],
    ("大阪全城", "芝士蛋糕/甜品"): [
        food_pin("Rikuro老爷爷难波本店", "Rikuro Ojisan no Mise Namba Main Store"),
    ],
    ("姬路", "当地午餐"): [
        food_pin("柊 穴子料理", "Anago Restaurant Hiiragi Himeji"),
        food_pin("姬路おでん能古", "Himeji Oden Nodaya"),
    ],
    ("神户三宫/元町", "神户牛"): [
        food_pin("Steakland Kobe-kan", "Steakland Kobe-kan"),
        food_pin("Mouriya本店", "Mouriya Honten Kobe"),
        food_pin("Wakkoqu新神户", "Wakkoqu Shin-Kobe"),
    ],
    ("神户南京町", "小吃"): [
        food_pin("老祥记", "Roushouki Kobe Nankinmachi"),
        food_pin("YUNYUN", "YUNYUN Kobe Nankinmachi"),
    ],
    ("神户港湾", "甜品/咖啡"): [
        food_pin("Kobe Harborland umie MOSAIC", "Kobe Harborland umie MOSAIC", "商场内可现场选择咖啡和甜品"),
    ],
    ("USJ", "园区餐"): [
        food_pin("Kinopio's Cafe", "Kinopios Cafe Universal Studios Japan"),
        food_pin("Three Broomsticks", "Three Broomsticks Universal Studios Japan"),
    ],
    ("京都祇园/东山", "京料理/天妇罗"): [
        food_pin("祇园天ぷら八坂圓堂", "Gion Tempura Endo Yasaka Kyoto"),
    ],
    ("京都先斗町/木屋町", "居酒屋/京味小菜"): [
        food_pin("Pontocho Robin", "Pontocho Robin Kyoto"),
    ],
    ("京都清水寺周边", "小吃甜品"): [
        food_pin("MACCHA HOUSE清水产宁坂", "MACCHA HOUSE Kyoto Kiyomizu Sannenzaka"),
        food_pin("伊藤久右卫门清水坂店", "Itoh Kyuemon Kiyomizu-saka Store"),
    ],
    ("京都岚山", "汤豆腐/荞麦"): [
        food_pin("松籁庵", "Shoraian Arashiyama Kyoto"),
        food_pin("岚山よしむら", "Arashiyama Yoshimura Kyoto"),
    ],
    ("京都岚山", "咖啡"): [
        food_pin("% ARABICA京都岚山", "% ARABICA Kyoto Arashiyama"),
    ],
    ("京都站", "拉面/车站餐"): [
        food_pin("京都拉面小路", "Kyoto Ramen Koji"),
        food_pin("Kyoto Porta Dining", "Kyoto Porta Dining"),
    ],
    ("宇治", "抹茶正餐/甜品"): [
        food_pin("中村藤吉本店", "Nakamura Tokichi Honten Uji"),
        food_pin("伊藤久右卫门宇治本店", "Itoh Kyuemon Uji Honten"),
        food_pin("通圆茶屋本店", "Tsuen Tea Uji"),
    ],
    ("宇治", "鳗鱼/京料理"): [
        food_pin("京・宇治 抹茶料理辰巳屋", "Kyo-Uji Shunsai Tatsumiya"),
    ],
    ("奈良", "小吃"): [
        food_pin("中谷堂", "Nakatanidou Nara"),
        food_pin("大佛布丁梦风广场店", "Mahoroba Daibutsu Purin Yumekaze Square Nara"),
    ],
    ("奈良", "正餐"): [
        food_pin("志津香公园店", "Shizuka Park Store Nara"),
        food_pin("平宗奈良店", "Hiraso Nara Store"),
    ],
    ("贵船/鞍马", "山里午餐"): [
        food_pin("贵船ひろ文", "Kibune Hirobun Kyoto"),
        food_pin("贵船伝兵衛", "Kibune Denbei Kyoto"),
    ],
    ("大原", "乡土料理"): [
        food_pin("志野 松门", "Shino Shoumon Ohara Kyoto"),
    ],
    ("伏见/京都市区", "清酒/拉面/市场"): [
        food_pin("鸟せい本店", "Torisei Honten Fushimi Kyoto"),
        food_pin("锦市场", "Nishiki Market Kyoto"),
    ],
    ("东京涩谷", "乌冬/夜景餐"): [
        food_pin("TsuruTonTan涩谷", "TsuruTonTan Shibuya Scramble Square"),
    ],
    ("东京原宿/表参道", "逛街午餐"): [
        food_pin("Maisen青山本店", "Tonkatsu Maisen Aoyama Honten"),
        food_pin("bills表参道", "bills Omotesando"),
    ],
    ("东京浅草", "老派早餐/甜品"): [
        food_pin("Ginza Brazil", "Ginza Brazil Asakusa"),
        food_pin("浅草Silk Pudding", "Asakusa Silk Pudding"),
    ],
    ("东京银座/东京站", "购物日餐饮"): [
        food_pin("根室花丸 KITTE丸之内", "Nemuro Hanamaru KITTE Marunouchi"),
        food_pin("六厘舍 东京拉面街", "Rokurinsha Tokyo Station"),
        food_pin("银座 篝本店", "Ginza Kagari Honten"),
    ],
    ("河口湖", "地方菜"): [
        food_pin("ほうとう不动 河口湖站前店", "Houtou Fudou Kawaguchiko Station"),
        food_pin("甲州ほうとう小作 河口湖店", "Kosaku Kawaguchiko"),
    ],
    ("镰仓/江之岛", "海边饭"): [
        food_pin("秋本", "Akimoto Kamakura"),
        food_pin("とびっちょ江之岛本店", "Tobiccho Enoshima"),
        food_pin("bills七里滨", "bills Shichirigahama"),
    ],
    ("机场/伴手礼", "伴手礼"): [
        food_pin("Tokyo Milk Cheese Factory羽田机场", "Tokyo Milk Cheese Factory Haneda Airport"),
    ],
}


# 新增条目优先补足原攻略较少的拉面、寿司、咖啡与东京单店选择。
EXTRA_FOOD_RECOMMENDATIONS: list[dict[str, object]] = [
    {
        "区域": "大阪福岛",
        "类型": "拉面",
        "推荐店/吃法": "燃えよ麺助：鸭肉/贝类系拉面，适合从梅田步行或坐一站去吃",
        "适合日期": "Day2/Day4",
        "预算感": "低-中",
        "预约/排队": "不能预约，常排队；开门前或午后错峰",
        "点单建议": "鸭出汁酱油拉面、贝类盐味拉面",
        "P人备选": "队伍太长就回梅田地下街吃拉面",
        "地图店铺": [food_pin("燃えよ麺助", "Moeyo Mensuke Osaka")],
    },
    {
        "区域": "大阪中央市场",
        "类型": "寿司",
        "推荐店/吃法": "Endo Sushi中央市场店：传统抓握寿司，适合愿意早起的人",
        "适合日期": "Day2",
        "预算感": "中",
        "预约/排队": "早市型店铺，出发前确认营业日与结束时间",
        "点单建议": "先点一盘综合寿司，不够再加",
        "P人备选": "不早起就改梅田春驹或寿司酒场",
        "地图店铺": [food_pin("Endo Sushi中央市场店", "Endo Sushi Central Market Osaka")],
    },
    {
        "区域": "大阪心斋桥",
        "类型": "精品咖啡",
        "推荐店/吃法": "LiLo Coffee Roasters：逛心斋桥时顺路补一杯",
        "适合日期": "Day1/Day2",
        "预算感": "低-中",
        "预约/排队": "无需预约，座位不多",
        "点单建议": "手冲或拿铁，豆子可作轻便伴手礼",
        "P人备选": "满座就外带继续逛",
        "地图店铺": [food_pin("LiLo Coffee Roasters", "LiLo Coffee Roasters Osaka")],
    },
    {
        "区域": "姬路站周边",
        "类型": "乌冬",
        "推荐店/吃法": "Menme：手打乌冬，适合作为姬路城往返途中的午餐",
        "适合日期": "Day3",
        "预算感": "低-中",
        "预约/排队": "店面不大，午餐时段可能排队",
        "点单建议": "冷乌冬更显筋道，天气凉则选热汤乌冬",
        "P人备选": "赶车就改姬路站内定食",
        "地图店铺": [food_pin("Menme", "Menme Himeji Udon")],
    },
    {
        "区域": "神户三宫",
        "类型": "神户洋食",
        "推荐店/吃法": "Grill Ippei：不吃高价神户牛时，可选蛋包饭、炸虾等老派洋食",
        "适合日期": "Day3",
        "预算感": "中",
        "预约/排队": "饭点可能等位，适合提前或延后吃",
        "点单建议": "蛋包饭、炸虾、汉堡排",
        "P人备选": "等位太久就去南京町少量多吃",
        "地图店铺": [food_pin("Grill Ippei三宫", "Grill Ippei Sannomiya Kobe")],
    },
    {
        "区域": "京都银阁寺",
        "类型": "乌冬",
        "推荐店/吃法": "Omen银阁寺本店：蔬菜配料丰富，口味清爽",
        "适合日期": "Day8自由调整",
        "预算感": "中",
        "预约/排队": "热门时段会排队，开门前后更稳",
        "点单建议": "招牌Omen乌冬",
        "P人备选": "不去银阁寺则无需专程跨城",
        "地图店铺": [food_pin("Omen银阁寺本店", "Omen Ginkakuji Kyoto")],
    },
    {
        "区域": "京都冈崎",
        "类型": "乌冬",
        "推荐店/吃法": "山元麺蔵：高人气手打乌冬，只适合愿意为一餐留时间的人",
        "适合日期": "Day5/Day8",
        "预算感": "中",
        "预约/排队": "规则可能调整，出发前看地图页最新公告",
        "点单建议": "牛蒡天妇罗乌冬",
        "P人备选": "不为网红店打乱当天东山动线",
        "地图店铺": [food_pin("山元麺蔵", "Yamamoto Menzou Kyoto")],
    },
    {
        "区域": "京都市中心",
        "类型": "荞麦",
        "推荐店/吃法": "本家尾张屋本店：京都老字号荞麦，适合想吃传统口味",
        "适合日期": "Day5/Day8",
        "预算感": "中",
        "预约/排队": "午餐可能排队，闭店较早",
        "点单建议": "宝来荞麦或天妇罗荞麦",
        "P人备选": "行程不顺路就选京都站或河原町分店型餐饮",
        "地图店铺": [food_pin("本家尾张屋本店", "Honke Owariya Honten Kyoto")],
    },
    {
        "区域": "京都市中心",
        "类型": "咖啡/早餐",
        "推荐店/吃法": "Inoda Coffee本店适合老派早餐；Weekenders适合快速喝精品咖啡",
        "适合日期": "Day5-Day9",
        "预算感": "低-中",
        "预约/排队": "Inoda早餐可能排队；Weekenders座位很少",
        "点单建议": "Inoda早餐套餐；Weekenders手冲/拿铁",
        "P人备选": "按当天动线二选一，不必都打卡",
        "地图店铺": [
            food_pin("Inoda Coffee本店", "Inoda Coffee Honten Kyoto"),
            food_pin("Weekenders Coffee富小路", "Weekenders Coffee Tominokoji Kyoto"),
        ],
    },
    {
        "区域": "京都四条乌丸",
        "类型": "创意寿司",
        "推荐店/吃法": "AWOMB乌丸本店：手织寿司摆盘漂亮，适合安排一顿仪式感午餐",
        "适合日期": "Day5/Day8",
        "预算感": "中-高",
        "预约/排队": "建议提前预约并确认当日套餐",
        "点单建议": "手织寿司套餐",
        "P人备选": "约不到就去锦市场/河原町吃普通寿司",
        "地图店铺": [food_pin("AWOMB乌丸本店", "AWOMB Karasuma Honten Kyoto")],
    },
    {
        "区域": "东京新宿",
        "类型": "乌冬",
        "推荐店/吃法": "Udon Shin：新宿热门手打乌冬，住新宿可较方便地错峰",
        "适合日期": "Day9-Day11",
        "预算感": "低-中",
        "预约/排队": "常排队，先看地图实时繁忙度",
        "点单建议": "培根天妇罗釜玉乌冬或冷乌冬",
        "P人备选": "超过可接受等待时间就回新宿站商场",
        "地图店铺": [food_pin("Udon Shin", "Udon Shin Shinjuku")],
    },
    {
        "区域": "东京新宿",
        "类型": "沾面",
        "推荐店/吃法": "风云儿：浓厚鸡白汤鱼介沾面，适合回酒店前快速吃",
        "适合日期": "Day9-Day11",
        "预算感": "低",
        "预约/排队": "翻台较快但高峰仍会排队",
        "点单建议": "特制沾面，食量小选普通份",
        "P人备选": "不想排就找车站内拉面",
        "地图店铺": [food_pin("风云儿", "Fuunji Shinjuku")],
    },
    {
        "区域": "东京涩谷",
        "类型": "平价寿司",
        "推荐店/吃法": "鱼べい道玄坂店：点单式寿司，购物途中补充体力方便",
        "适合日期": "Day9/Day10",
        "预算感": "低-中",
        "预约/排队": "高峰可能排队，现场取号",
        "点单建议": "少量多轮点，避免一次点太多",
        "P人备选": "队伍长就进涩谷商场餐饮层",
        "地图店铺": [food_pin("鱼べい涩谷道玄坂店", "Uobei Shibuya Dogenzaka")],
    },
    {
        "区域": "东京原宿",
        "类型": "柚子拉面",
        "推荐店/吃法": "AFURI原宿：柚子盐拉面清爽，适合逛街日",
        "适合日期": "Day10",
        "预算感": "低-中",
        "预约/排队": "不能预约，翻台通常较快",
        "点单建议": "柚子盐拉面，想浓一点可选柚子辣露",
        "P人备选": "表参道人多时可留到涩谷再吃",
        "地图店铺": [food_pin("AFURI原宿", "AFURI Harajuku")],
    },
    {
        "区域": "东京浅草",
        "类型": "饭团早餐",
        "推荐店/吃法": "浅草宿六：老牌现做饭团，适合回程日上午吃轻早餐",
        "适合日期": "Day11/Day12",
        "预算感": "低-中",
        "预约/排队": "座位少，营业时段需出发前确认",
        "点单建议": "两枚饭团加味噌汤",
        "P人备选": "回程日遇到排队直接跳过",
        "地图店铺": [food_pin("浅草宿六", "Onigiri Asakusa Yadoroku")],
    },
    {
        "区域": "东京浅草",
        "类型": "铁板烧",
        "推荐店/吃法": "染太郎：日式老屋中的文字烧/大阪烧体验",
        "适合日期": "Day11东京自由",
        "预算感": "中",
        "预约/排队": "用餐时间较长，不建议赶飞机当天去",
        "点单建议": "文字烧与大阪烧各点一种分食",
        "P人备选": "时间紧就吃浅草小吃",
        "地图店铺": [food_pin("染太郎", "Sometaro Asakusa")],
    },
    {
        "区域": "东京日本桥",
        "类型": "海鲜丼",
        "推荐店/吃法": "つじ半日本桥本店：海鲜珠宝丼，最后可加高汤做茶泡饭",
        "适合日期": "Day11东京自由/Day12",
        "预算感": "中",
        "预约/排队": "热门时段排队明显，不适合赶飞机前硬等",
        "点单建议": "基础梅套餐已足够丰富",
        "P人备选": "排队长就改东京站商场寿司",
        "地图店铺": [food_pin("つじ半日本桥本店", "Tsujihan Nihonbashi")],
    },
    {
        "区域": "镰仓由比滨",
        "类型": "荞麦",
        "推荐店/吃法": "松原庵：古民家荞麦与天妇罗，适合镰仓慢游",
        "适合日期": "Day11镰仓",
        "预算感": "中-高",
        "预约/排队": "建议预约；用餐节奏较慢",
        "点单建议": "荞麦加天妇罗套餐",
        "P人备选": "时间紧就小町通或江之岛吃吻仔鱼丼",
        "地图店铺": [food_pin("镰仓松原庵", "Kamakura Matsubaraan")],
    },
]


def hotel_pick(
    name: str,
    area: str,
    budget: str,
    fit: str,
    query: str,
) -> dict[str, str]:
    """保存住宿选择所需的区域、预算定位和地图目标，便于用户从当前位置直接导航。"""
    return {
        "名称": name,
        "区域": area,
        "预算感": budget,
        "适合理由": fit,
        "地图查询": query,
    }


HOTEL_RECOMMENDATIONS: dict[str, list[dict[str, str]]] = {
    "大阪段": [
        hotel_pick(
            "东横INN大阪难波西",
            "难波/元町",
            "约¥350–500/间夜",
            "独立卫浴，通常含简易早餐；步行可到南海难波站，兼顾关西机场与道顿堀。",
            "Toyoko Inn Osaka Namba Nishi",
        ),
        hotel_pick(
            "东横INN大阪难波日本桥",
            "日本桥/难波",
            "约¥350–500/间夜",
            "独立卫浴并含简易早餐，靠近黑门市场；去难波、心斋桥和机场都方便。",
            "Toyoko Inn Osaka Namba Nippombashi",
        ),
        hotel_pick(
            "KOKO HOTEL 大阪难波惠美须町",
            "惠美须町/新世界",
            "约¥300–500/间夜",
            "独立卫浴；地铁堺筋线出行方便，房价通常比难波核心区友好。",
            "KOKO HOTEL Osaka Namba Ebisucho",
        ),
        hotel_pick(
            "KOKO HOTEL 大阪梅田",
            "梅田/大阪站东侧",
            "促销约¥400–500/间夜",
            "独立卫浴；到大阪站约步行10分钟，去姬路、神户和USJ换乘顺，超预算时跳过。",
            "KOKO HOTEL Osaka Umeda",
        ),
        hotel_pick(
            "Hotel Sunplaza 2 Annex",
            "新今宫/动物园前",
            "约¥200–400/间夜",
            "经济型私人客房，部分房型使用共用浴场/卫浴；JR、南海和地铁三线可用。",
            "Hotel Sunplaza 2 Annex Osaka",
        ),
        hotel_pick(
            "Hotel Toyo Osaka",
            "动物园前/新今宫",
            "约¥150–300/间夜",
            "青旅型私人房、共用卫浴，适合单人压低预算；在意隔音和卫浴私密性则不选。",
            "Hotel Toyo Osaka",
        ),
    ],
    "京都段": [
        hotel_pick(
            "THE POCKET HOTEL 京都乌丸五条",
            "五条/乌丸线",
            "约¥250–450/间夜",
            "带锁私人房、共用卫浴；五条站步行约1分钟，适合单人住且交通很省力。",
            "THE POCKET HOTEL Kyoto Karasuma Gojo",
        ),
        hotel_pick(
            "HOTEL TAVINOS Kyoto",
            "河原町五条/清水五条",
            "约¥300–500/间夜",
            "紧凑型独立客房，步行可到清水五条站；去祇园、清水寺与京都站较均衡。",
            "HOTEL TAVINOS Kyoto",
        ),
        hotel_pick(
            "ibis Styles Kyoto Shijo",
            "四条乌丸",
            "约¥350–500/间夜",
            "独立卫浴，四条站和乌丸站步行约4分钟；位置贴合原攻略，促销价合适时优先。",
            "ibis Styles Kyoto Shijo",
        ),
        hotel_pick(
            "HOTEL M's PLUS SHIJO OMIYA",
            "四条大宫",
            "约¥300–500/间夜",
            "独立卫浴，靠近阪急与岚电；去岚山方便，坐一小段车即可到四条河原町。",
            "HOTEL M's PLUS SHIJO OMIYA",
        ),
        hotel_pick(
            "WeBase 京都",
            "四条乌丸",
            "约¥200–450/人夜",
            "青旅/简约酒店，按胶囊床位或私人房选择；四条站步行约5分钟，适合单人。",
            "WeBase Kyoto",
        ),
        hotel_pick(
            "东横INN京都四条大宫",
            "四条大宫",
            "约¥350–500/间夜",
            "独立卫浴并通常含简易早餐，适合看重稳定连锁标准和交通便利的人。",
            "Toyoko Inn Kyoto Shijo Omiya",
        ),
    ],
    "东京段": [
        hotel_pick(
            "HOTEL TAVINOS Asakusa",
            "浅草",
            "约¥350–500/间夜",
            "紧凑型独立客房，适合浅草与上野动线；去新宿约需半小时，优先保预算。",
            "HOTEL TAVINOS Asakusa",
        ),
        hotel_pick(
            "Agora Place Tokyo Asakusa",
            "田原町/浅草",
            "约¥400–500/间夜",
            "独立卫浴，田原町站步行约1分钟；银座线上野、银座和涩谷无需换乘。",
            "Agora Place Tokyo Asakusa",
        ),
        hotel_pick(
            "HOTEL MYSTAYS Asakusa",
            "本所/藏前",
            "约¥350–500/间夜",
            "独立卫浴并有投币洗衣房；位置不在最热闹街区，通常更容易守住预算。",
            "HOTEL MYSTAYS Asakusa",
        ),
        hotel_pick(
            "Ueno New Izu Hotel",
            "上野/稻荷町",
            "约¥350–500/间夜",
            "老牌经济型酒店，步行可到上野站；适合成田机场、浅草和东京站方向。",
            "Ueno New Izu Hotel Tokyo",
        ),
        hotel_pick(
            "Juyoh Hotel",
            "南千住/山谷",
            "约¥200–350/间夜",
            "经济型私人房、共用卫浴；预算优先的单人备选，夜间回程需多留意周边环境。",
            "Juyoh Hotel Tokyo",
        ),
        hotel_pick(
            "Hotel Palace Japan",
            "南千住",
            "约¥200–400/间夜",
            "带锁私人房、共用卫浴，价格通常低于核心区商务酒店；适合单人轻装入住。",
            "Hotel Palace Japan Tokyo",
        ),
        hotel_pick(
            "9h nine hours Shinjuku-North",
            "新大久保/新宿北侧",
            "约¥250–450/人夜",
            "胶囊房，2026年开业；想保留新宿动线且能接受共用卫浴时再选。",
            "9h nine hours Shinjuku-North",
        ),
    ],
    "可选温泉": [
        hotel_pick(
            "东横INN神户三宫1号店",
            "神户三宫",
            "约¥350–500/间夜",
            "独立卫浴并通常含简易早餐；住三宫后白天往返有马温泉，比住温泉旅馆省钱。",
            "Toyoko Inn Kobe Sannomiya No.1",
        ),
        hotel_pick(
            "Kobe Sannomiya Union Hotel",
            "神户三宫东侧",
            "约¥300–500/间夜",
            "独立卫浴的商务酒店，适合把神户作为有马温泉日归基地。",
            "Kobe Sannomiya Union Hotel",
        ),
        hotel_pick(
            "Kobe City Gardens Hotel",
            "JR神户站",
            "约¥250–450/间夜",
            "经济型独立客房，靠近JR神户站；适合先看港区、再搭车去有马。",
            "Kobe City Gardens Hotel",
        ),
        hotel_pick(
            "Hostel Anchorage",
            "神户站/港区",
            "约¥180–350/人夜",
            "青旅床位或简约私人房、共用卫浴；预算最优先时选择。",
            "Hostel Anchorage Kobe",
        ),
    ],
}


# USJ 官方不发布未来逐项目排队时间。项目说明、活动日期、休止与票券规则以官方为准；
# 排队区间结合 2026 年历史等待数据和 9 月客流预测，专门服务本行程的 9 月 27—29 日决策。
USJ_GUIDE: dict[str, object] = {
    "summary": {
        "结论": "客流优先建议把USJ调整到9月27日，并购买同时覆盖咚奇刚与马里奥赛车的Express Pass 4；9月28日改走姬路＋神户，酒店无需变更。",
        "预测说明": "表内排队时间是天气正常、项目正常运行时10:00—17:00的常见区间，不是USJ官方承诺；开园直冲、单人通道和临近闭园可能更短。",
        "当天动作": "前一晚把所有人的入园票绑定官方App；无任天堂入场确约券时，过闸后立即申请超级任天堂世界e整理券。",
    },
    "dates": [
        {
            "日期": "9月27日 周日",
            "营业时间": "预计09:00—21:30",
            "预测平均等待": "约53分钟",
            "年卡情况": "标准年卡除外日",
            "行程影响": "与28日姬路＋神户互换即可；大阪住宿无需调整",
            "建议": "客流首选",
        },
        {
            "日期": "9月28日 周一",
            "营业时间": "预计09:00—21:30",
            "预测平均等待": "约65分钟",
            "年卡情况": "标准年卡除外日",
            "行程影响": "继续住大阪，轻装往返；整体最顺",
            "建议": "原计划备选",
        },
        {
            "日期": "9月29日 周二",
            "营业时间": "预计08:00—22:00",
            "预测平均等待": "约67分钟",
            "年卡情况": "标准年卡恢复入园",
            "行程影响": "原计划大阪退房转京都；改期需处理行李并牺牲东山行程",
            "建议": "只有延住大阪一晚才考虑",
        },
    ],
    "attractions": [
        {
            "名称": "咚奇刚的疯狂矿车",
            "区域": "超级任天堂世界",
            "类型": "过山车／任天堂",
            "介绍": "以矿车跳轨错觉和丛林场景为核心，刺激度中等，但新鲜感和主题完成度很高。",
            "推荐度": 5,
            "预计排队": "150—240分钟",
            "速通建议": "全园第一速通目标；若不买速通，开园后直接冲",
            "单人通道": "官方列有单人通道，当日可能暂停",
            "取舍": "必玩",
        },
        {
            "名称": "马里奥赛车～库巴的挑战书～",
            "区域": "超级任天堂世界",
            "类型": "AR互动乘车／任天堂",
            "介绍": "佩戴AR设备在库巴城堡中射击龟壳，排队区本身也是完整布景体验。",
            "推荐度": 5,
            "预计排队": "100—160分钟",
            "速通建议": "与咚奇刚一起锁定；第一次体验建议走普通队看完整布景",
            "单人通道": "官方列有单人通道，但会拆散同行者",
            "取舍": "必玩",
        },
        {
            "名称": "耀西冒险",
            "区域": "超级任天堂世界",
            "类型": "景观型乘车",
            "介绍": "节奏平缓，主要价值是从高处观看任天堂世界和寻找彩蛋。",
            "推荐度": 3,
            "预计排队": "60—100分钟",
            "速通建议": "不值得单独为它买速通；排队超过70分钟可放弃",
            "单人通道": "无",
            "取舍": "任天堂粉丝或拍照党保留",
        },
        {
            "名称": "哈利·波特与禁忌之旅",
            "区域": "哈利·波特魔法世界",
            "类型": "室内动感乘车",
            "介绍": "在霍格沃茨城堡内结合机械臂、实景与影像飞行，沉浸感很强。",
            "推荐度": 5,
            "预计排队": "50—90分钟",
            "速通建议": "Express Pass 4的优质第四项；不买时可留意单人通道",
            "单人通道": "官方列有单人通道",
            "取舍": "首次到访必玩；容易晕3D者谨慎",
        },
        {
            "名称": "鹰马的飞行",
            "区域": "哈利·波特魔法世界",
            "类型": "家庭过山车",
            "介绍": "短程户外过山车，能看到海格小屋，刺激度较低。",
            "推荐度": 2,
            "预计排队": "60—100分钟",
            "速通建议": "项目很短，普通队超过45分钟就跳过",
            "单人通道": "无",
            "取舍": "时间充裕再玩",
        },
        {
            "名称": "霍格沃茨城堡漫步",
            "区域": "哈利·波特魔法世界",
            "类型": "步行参观／期间限定",
            "介绍": "沿特别路线参观画像大厅等城堡细节，2026年9月1日至2027年1月17日回归。",
            "推荐度": 4,
            "预计排队": "10—30分钟",
            "速通建议": "无需为此购买常规速通",
            "单人通道": "不适用",
            "取舍": "哈迷推荐，也适合当作休整段",
        },
        {
            "名称": "飞天翼龙",
            "区域": "侏罗纪公园",
            "类型": "高强度过山车",
            "介绍": "俯卧悬挂式高速过山车，失重和翻转明显，是园内刺激项目代表。",
            "推荐度": 5,
            "预计排队": "60—100分钟",
            "速通建议": "刺激党优先；不买时使用官方单人通道",
            "单人通道": "官方列有单人通道",
            "取舍": "刺激党必玩；颈背或脚踝仍不适者先确认乘坐限制",
        },
        {
            "名称": "好莱坞美梦乘车游",
            "区域": "好莱坞",
            "类型": "音乐过山车／正向",
            "介绍": "可选择座椅音乐的户外过山车，速度感强，整体比飞天翼龙温和。",
            "推荐度": 4,
            "预计排队": "70—110分钟",
            "速通建议": "可用官方单人通道；与逆转世界二选一",
            "单人通道": "官方仅正向版本提供，逆向不可用",
            "取舍": "喜欢过山车就保留",
        },
        {
            "名称": "好莱坞美梦～逆转世界～",
            "区域": "好莱坞",
            "类型": "音乐过山车／倒退",
            "介绍": "同一轨道倒退运行，未知感更强，队伍通常比正向更长。",
            "推荐度": 4,
            "预计排队": "80—130分钟",
            "速通建议": "没有单人通道；特别想玩再选含它的套餐",
            "单人通道": "无",
            "取舍": "正向和逆向只选一个即可",
        },
        {
            "名称": "侏罗纪公园乘船游～暗夜版～",
            "区域": "侏罗纪公园",
            "类型": "激流勇进／万圣节限定",
            "介绍": "白天为经典恐龙乘船项目，万圣节夜间加入全黑与惊吓演出。",
            "推荐度": 4,
            "预计排队": "白天30—60分钟；夜间50—90分钟",
            "速通建议": "普通队通常可接受；想看暗夜版需留到晚上",
            "单人通道": "官方列有单人通道，当日以现场为准",
            "取舍": "夜间优先，准备雨衣或接受湿身",
        },
        {
            "名称": "大白鲨～Red Alert～",
            "区域": "亲善村",
            "类型": "乘船演出／万圣节限定",
            "介绍": "由船长真人演出推进剧情；夜间红光、雾气和低能见度版本氛围更好。",
            "推荐度": 4,
            "预计排队": "30—60分钟",
            "速通建议": "通常不必占用核心速通名额，晚间排普通队",
            "单人通道": "官方列有单人通道",
            "取舍": "建议夜间体验",
        },
        {
            "名称": "小黄人疯狂乘车游",
            "区域": "小黄人乐园",
            "类型": "室内模拟器",
            "介绍": "以巨大屏幕模拟高速移动，亲子友好，但视觉晃动明显。",
            "推荐度": 3,
            "预计排队": "45—80分钟",
            "速通建议": "非小黄人粉丝不必占用高价套餐名额",
            "单人通道": "官方列有单人通道",
            "取舍": "容易晕动可跳过",
        },
        {
            "名称": "小黄人大恶党任务",
            "区域": "小黄人乐园",
            "类型": "互动射击",
            "介绍": "边移动边完成射击任务，刺激度低，适合作为轻松项目。",
            "推荐度": 3,
            "预计排队": "30—60分钟",
            "速通建议": "排队短时再玩",
            "单人通道": "无",
            "取舍": "亲子或小黄人粉丝推荐",
        },
        {
            "名称": "名侦探柯南4-D现场秀",
            "区域": "纽约／剧场",
            "类型": "4-D演出",
            "介绍": "真人表演结合影像和座椅效果，日语对白较多。",
            "推荐度": 3,
            "预计排队": "30—60分钟",
            "速通建议": "动漫粉丝再考虑，具体演出以当日时刻表为准",
            "单人通道": "不适用",
            "取舍": "听不懂日语且时间紧可跳过",
        },
        {
            "名称": "水世界",
            "区域": "水世界",
            "类型": "大型特技秀",
            "介绍": "包含水上追逐、爆破和飞机特技，是USJ最值得看的固定演出之一。",
            "推荐度": 5,
            "预计排队": "开演前20—35分钟到场",
            "速通建议": "按演出时刻安排，不需要常规速通",
            "单人通道": "不适用",
            "取舍": "建议用作下午坐下休息",
        },
        {
            "名称": "SING ON TOUR",
            "区域": "好莱坞／剧场",
            "类型": "室内音乐秀",
            "介绍": "真人和大型角色同台演唱，室内有座位，适合恢复体力。",
            "推荐度": 4,
            "预计排队": "开演前10—25分钟到场",
            "速通建议": "无需速通",
            "单人通道": "不适用",
            "取舍": "中午或下午安排一场",
        },
    ],
    "halloween": [
        {
            "名称": "街头僵尸＋Zombie de Dance",
            "时间": "通常18:00至闭园",
            "预计等待": "无需固定排队；热门表演点需提前占位",
            "建议": "至少留60分钟，9月27—29日都有普通万圣惊魂夜",
        },
        {
            "名称": "Factory of Fear 僵尸工厂",
            "时间": "官方公布10:00至闭园",
            "预计等待": "整理券制时按指定时间；普通队约60—120分钟",
            "建议": "可能需要App e整理券，且14岁以下不可体验",
        },
        {
            "名称": "生化危机 Requiem：The Dive",
            "时间": "以当日官方时刻表为准",
            "预计等待": "约60—120分钟或指定时段",
            "建议": "恐怖爱好者优先；可比较万圣节限定Express Pass 4",
        },
        {
            "名称": "电锯人 The Chaos 4-D",
            "时间": "9月11日至11月8日期间限定",
            "预计等待": "约40—80分钟",
            "建议": "动漫粉丝推荐，日语对白会影响部分理解",
        },
        {
            "名称": "贞子的诅咒：暗黑恐怖乘车",
            "时间": "以当日官方时刻表为准",
            "预计等待": "约60—100分钟",
            "建议": "普通Space Fantasy长期休止，但万圣节期间使用该设施运行贞子主题版本",
        },
        {
            "名称": "18号宅邸的魔女",
            "时间": "9月11日至11月8日期间限定",
            "预计等待": "约45—90分钟或整理券制",
            "建议": "喜欢步行式鬼屋再选；先处理任天堂和核心项目",
        },
    ],
    "express_pass4": [
        {
            "名称": "矿车＆欢乐（トロッコ＆ファン）",
            "标签": "均衡首选",
            "项目": [
                "咚奇刚的疯狂矿车",
                "马里奥赛车～库巴的挑战书～",
                "哈利·波特与禁忌之旅",
                "大白鲨／侏罗纪公园乘船游（二选一）",
            ],
            "推荐理由": "第一次去最稳妥：同时保住任天堂两大热门项目，并覆盖哈利·波特；通常还附超级任天堂世界和哈利·波特区域入场保证。",
            "情侣建议": "咚奇刚、马里奥和哈利·波特用速通一起乘坐；第四项想体验万圣节夜间版本优先选侏罗纪，怕湿或脚踝不适则选大白鲨。",
        },
        {
            "名称": "矿车＆飞天翼龙（トロッコ＆フライングダイナソー）",
            "标签": "刺激党",
            "项目": [
                "咚奇刚的疯狂矿车",
                "马里奥赛车～库巴的挑战书～",
                "飞天翼龙",
                "大白鲨／侏罗纪公园乘船游（二选一）",
            ],
            "推荐理由": "两个人都喜欢过山车时，纯省时价值更高；哈利·波特禁忌之旅可以另外使用单人通道。",
            "情侣建议": "飞天翼龙强度高；颈背、脚踝仍不舒服或容易晕车时不要为了套餐勉强乘坐。",
        },
        {
            "名称": "万圣节限定 Express Pass 4",
            "标签": "恐怖主题",
            "项目": [
                "以购买页显示的生化危机／电锯人等限定项目为准",
                "部分组合还会搭配贞子、侏罗纪暗夜版或好莱坞美梦",
            ],
            "推荐理由": "只有把万圣节鬼屋和限定项目放在第一优先级时再选；它通常不能同时解决咚奇刚和马里奥的长队。",
            "情侣建议": "第一次到USJ仍优先购买常规矿车组合；街头僵尸和Zombie de Dance不需要购买速通。",
        },
    ],
    "single_rider_couples": [
        {
            "项目": "飞天翼龙",
            "建议": "强烈推荐",
            "情侣策略": "重点是刺激体验，分开乘坐影响较小；不在速通套餐中时优先走单人通道。",
        },
        {
            "项目": "好莱坞美梦乘车游（正向）",
            "建议": "推荐",
            "情侣策略": "乘坐时本来也很难交流，适合用单人通道省时；逆转世界没有单人通道。",
        },
        {
            "项目": "哈利·波特与禁忌之旅",
            "建议": "推荐",
            "情侣策略": "速通未包含时很实用；如果两人都是哈迷，首次也可以一起排队看完整城堡布景。",
        },
        {
            "项目": "小黄人疯狂乘车游",
            "建议": "看现场时长",
            "情侣策略": "只有单人等待不超过普通队一半时才值得拆开。",
        },
        {
            "项目": "大白鲨／侏罗纪公园乘船游",
            "建议": "一般",
            "情侣策略": "一起坐的氛围、反应和合照价值更高；普通队不长时不建议拆开。",
        },
        {
            "项目": "咚奇刚的疯狂矿车",
            "建议": "首次不建议",
            "情侣策略": "两人一排的共同体验价值高，优先用速通一起坐；二刷再考虑单人通道。",
        },
        {
            "项目": "马里奥赛车～库巴的挑战书～",
            "建议": "首次不建议",
            "情侣策略": "可以一起计分和互动，第一次用速通一起玩；二刷再用单人通道。",
        },
    ],
    "single_rider_rules": [
        "单人通道免费，但不是速通，只是用单人游客填补空座；当天可能暂停，也不保证一定更快。",
        "两个人可以一起进入单人队列，但必须接受不同排、不同车，甚至相隔数班，不能到分组处再要求坐在一起。",
        "当天看官方App或入口牌：单人等待不超过普通队一半时直接用；只少10—20分钟时建议一起排。",
        "速通票每人都需要一张；两张应一次下单并选择相同时间段，避免被分到不同体验时段。",
    ],
    "route_assumption": {
        "适用日期": "2026年9月27日（周日）",
        "目标套餐": "Express Pass 4「矿车＆欢乐」",
        "建议时段": "优先选择任天堂世界10:00—12:30、哈利·波特15:00—17:00的组合",
        "固定节点": "咚奇刚、马里奥、哈利·波特及区域入场时间必须服从票面；其余项目按官方App实时等待调整。",
        "说明": "9月27日详细演出时刻尚未公布，时间轴以09:00官方开园、可能提前放行为基础；水世界、SING和柯南需在出发前一晚替换成官方实际场次。",
    },
    "route_timeline": [
        {
            "序号": 1,
            "时间": "07:15—08:00",
            "区域": "入口",
            "安排": "安检排队＋提前开园准备",
            "预计等待": "约45分钟，属于入园准备",
            "游玩占用": "—",
            "通道": "入园票",
            "理由": "即使官网写09:00开园，也给可能提前放行和安检留出余量；前一晚登录App并截好两人的票券。",
            "调整": "08:15仍未放行不影响后续，压缩小黄人项目即可。",
        },
        {
            "序号": 2,
            "时间": "开园后—09:00",
            "区域": "好莱坞",
            "安排": "好莱坞美梦乘车游（正向）",
            "预计等待": "普通队约20—45分钟；单人队通常更短",
            "游玩占用": "乘坐约3分钟；含存包、分组约15分钟",
            "通道": "单人／普通",
            "理由": "入口最近且上午队伍增长快；情侣愿意拆开就走单人，否则趁开园一起坐。",
            "调整": "显示超过55分钟就先跳过，闭园前再看。",
        },
        {
            "序号": 3,
            "时间": "09:05—09:55",
            "区域": "小黄人乐园",
            "安排": "小黄人大恶党任务",
            "预计等待": "约20—40分钟",
            "游玩占用": "体验约7分钟；总占用约35—50分钟",
            "通道": "普通队",
            "理由": "沿逆时针路线前进且上午通常比午后短；先在App内打开Universal Play。",
            "调整": "超过45分钟直接去任天堂世界，下午再补。",
        },
        {
            "序号": 4,
            "时间": "10:00—12:15",
            "区域": "超级任天堂世界",
            "安排": "按票面依次玩咚奇刚＋马里奥；空档拍照，耀西低于55分钟再排",
            "预计等待": "速通各约10—25分钟；耀西约45—80分钟",
            "游玩占用": "咚奇刚约2分钟、马里奥约5分钟；区域共约2小时",
            "通道": "Express Pass 4",
            "理由": "这是全天最难补救的固定节点；两人用速通一起乘坐，不用单人通道。",
            "调整": "任天堂入场若晚于11:30，把上午第3项移到17:00以后。",
        },
        {
            "序号": 5,
            "时间": "12:15—13:00",
            "区域": "侏罗纪公园／旧金山",
            "安排": "错峰午餐＋补水",
            "预计等待": "餐厅约10—25分钟",
            "游玩占用": "约40—45分钟",
            "通道": "—",
            "理由": "12点出头先吃，避开13点后的用餐高峰；晚上还有长时间站立活动。",
            "调整": "餐厅排队超过30分钟，改买移动餐饮或回到旧金山区域用餐。",
        },
        {
            "序号": 6,
            "时间": "13:00—14:05",
            "区域": "侏罗纪公园",
            "安排": "飞天翼龙",
            "预计等待": "单人约25—50分钟；普通队约60—100分钟",
            "游玩占用": "乘坐约3分钟；含存包、安检约20分钟",
            "通道": "情侣建议单人",
            "理由": "与午餐区相邻，且飞天翼龙分开乘坐的体验损失较小。",
            "调整": "单人超过55分钟就移到16:30后；脚踝、颈背不适直接取消。",
        },
        {
            "序号": 7,
            "时间": "14:10—15:00",
            "区域": "水世界",
            "安排": "观看最接近14:30的一场水世界",
            "预计等待": "开演前25—30分钟入场",
            "游玩占用": "演出约20分钟",
            "通道": "按演出时刻",
            "理由": "从侏罗纪步行很近；用坐席恢复体力，并避开下午排队峰值。",
            "调整": "最终场次不同就前后平移；湿身区通常有明显标识。",
        },
        {
            "序号": 8,
            "时间": "15:05—15:45",
            "区域": "亲善村",
            "安排": "大白鲨；若晚上想看Red Alert，此时只拍照并继续前进",
            "预计等待": "普通／单人约20—45分钟",
            "游玩占用": "乘船约7分钟；总占用约30—45分钟",
            "通道": "普通／单人",
            "理由": "位于水世界到哈利·波特的顺路位置；夜间版本18:30后才开始。",
            "调整": "准备晚上二刷时，下午这段改成黄油啤酒或提前进哈利区。",
        },
        {
            "序号": 9,
            "时间": "15:45—17:15",
            "区域": "哈利·波特魔法世界",
            "安排": "按票面玩禁忌之旅＋城堡漫步；鹰马低于45分钟再排",
            "预计等待": "速通约10—25分钟；城堡漫步约10—30分钟",
            "游玩占用": "禁忌之旅约5分钟；区域约75—90分钟",
            "通道": "Express Pass 4",
            "理由": "下午固定时段承接北侧路线，避免在园区两端来回折返。",
            "调整": "严格以票面时间为准；如果票面更早，把第7、8项整体后移。",
        },
        {
            "序号": 10,
            "时间": "17:20—18:20",
            "区域": "好莱坞／纽约",
            "安排": "SING或柯南4-D二选一＋简餐",
            "预计等待": "开场前15—25分钟",
            "游玩占用": "SING约20分钟；柯南约30分钟",
            "通道": "按演出时刻",
            "理由": "安排坐席休整，为万圣节夜间段保存体力；日语理解有限时优先SING。",
            "调整": "若场次不衔接，改补小黄人或商店购物。",
        },
        {
            "序号": 11,
            "时间": "18:30—21:15",
            "区域": "侏罗纪／亲善村／街区",
            "安排": "侏罗纪暗夜版（速通第四项）＋大白鲨Red Alert＋街头僵尸",
            "预计等待": "暗夜速通约10—30分钟；大白鲨约30—60分钟",
            "游玩占用": "夜间项目＋表演约2.5小时",
            "通道": "Express Pass 4＋普通队",
            "理由": "9月27日18:30后才进入两个暗夜版本；最后把时间留给万圣节氛围，不再安排远距离必玩项目。",
            "调整": "若第四项票券不能用于暗夜时段，白天使用，并把晚上重点改为鬼屋＋街头僵尸。",
        },
    ],
    "strategies": [
        {
            "方案": "完全不买速通",
            "追加费用": "¥0",
            "预计成果": "6—8个项目＋1场演出＋万圣节街头活动",
            "优点": "最省钱、路线最灵活，可充分利用单人通道。",
            "缺点": "需提前约1.5小时到门口；热门项目总排队可能6—9小时；任天堂入场存在风险。",
            "适合": "预算优先、愿意早起、能接受少玩几个项目的人。",
        },
        {
            "方案": "核心项目 Express Pass 4",
            "追加费用": "官方最低¥6,800起，9月底实际价以购票页为准",
            "预计成果": "8—11个项目＋1—2场演出＋万圣节",
            "优点": "通常可节省4—6小时，并锁定套餐包含的区域入场时段。",
            "缺点": "马里奥、咚奇刚等常规项目通常不能自由单买一个；套餐名称相似，必须展开核对4个具体项目。",
            "适合": "本行程首选；优先找同时含咚奇刚和马里奥赛车的组合。",
        },
        {
            "方案": "Express Pass 7／8",
            "追加费用": "官方最低EP7 ¥14,100起、EP8 ¥15,500起",
            "预计成果": "10—14个项目＋演出和夜间活动",
            "优点": "热门项目保障更完整，能把晚上留给万圣节，明显减少久站。",
            "缺点": "价格高且指定时段会限制路线；部分万圣节鬼屋仍可能需要单独整理券或限定快通。",
            "适合": "首次且短期内只去一次、脚踝不耐久站或不想早起的人。",
        },
        {
            "方案": "Premium 全部可用项目",
            "追加费用": "官方常规票种最低¥27,600起；实时商店价格可能更高",
            "预计成果": "最大限度压缩常规项目等待",
            "优点": "常规热门项目最省心。",
            "缺点": "性价比最低，也不等于所有万圣节限定项目全包；一天体力未必能用足价值。",
            "适合": "预算不敏感、唯一目标是尽可能少排队的人。",
        },
    ],
    "closures": [
        "2026年9月官方休止清单显示：Shrek 4-D、Sesame Street 4-D、普通Space Fantasy长期休止。",
        "普通Space Fantasy休止不等于贞子主题暗黑乘车取消；万圣节限定版本需看当天时刻表。",
        "天气、技术故障会临时停运；排队截止可能早于闭园，最后一小时不要押唯一必玩项目。",
    ],
    "sources": [
        {"名称": "USJ项目列表", "网址": "https://www.usj.co.jp/web/ja/jp/attractions"},
        {"名称": "USJ官方园区地图", "网址": "https://www.usj.co.jp/web/ja/jp/service-guide/parkmap"},
        {"名称": "USJ官方App（地图、等待时间与演出时刻）", "网址": "https://www.usj.co.jp/web/ja/jp/enjoy/app"},
        {"名称": "USJ营业时间", "网址": "https://www.usj.co.jp/web/ja/jp/park-guide/schedule/park-hour/"},
        {"名称": "USJ休止信息", "网址": "https://www.usj.co.jp/web/ja/jp/park-guide/schedule/attraction-closure"},
        {"名称": "USJ票种与最低价格", "网址": "https://www.usj.co.jp/web/ja/jp/tickets/lineup"},
        {"名称": "Express Pass实时组合", "网址": "https://prd-origin.usj.co.jp/ticket/expresspass/"},
        {"名称": "官方单人通道列表", "网址": "https://www.usj.co.jp/web/ja/jp/attractions/how-to-fun/single-rider"},
        {"名称": "2026万圣惊魂夜", "网址": "https://www.usj.co.jp/web/ja/jp/events/halloween-extreme-autumn-2026/halloween-horror-nights"},
        {"名称": "2026历史等待数据", "网址": "https://queue-times.com/en-US/parks/284/stats/2026?hide_archived=true"},
        {"名称": "2026年9月客流预测", "网址": "https://usjreal.asumirai.info/monthly/usj-forecast-2026-9.html"},
    ],
}


# 项目卡片将官网客观参数与本行程的实操建议分开，避免把估算时间误写成官方承诺。
USJ_ATTRACTION_DETAILS: dict[str, dict[str, str]] = {
    "咚奇刚的疯狂矿车": {
        "图片": "donkey-kong.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/attractions/donkey-kong-country-ride",
        "玩法": "乘坐2排矿车穿过黄金神殿和丛林，体验被发射、急转、水花以及“跳过断轨”的视觉错觉。它不是互动计分项目，重点是观察轨道机关、场景和彩蛋，并享受两人一排共同尖叫的过程。",
        "操作诀窍": "入座后背部贴紧靠背、脚放稳，松肩比全身绷紧更舒服；第一次和同行者一起坐，二刷才考虑单人通道。",
        "官方参数": "约2分钟｜4名（2名×2排）｜单独122cm以上；有陪同者107cm以上｜可能被水溅湿",
        "实际占用": "速通通常约15—30分钟；普通队需按App显示，热门时可能超过2小时。",
        "提醒": "有急加速、旋转和颠簸；脚踝、颈背或近期受伤仍不稳定时，先向现场工作人员确认。",
    },
    "马里奥赛车～库巴的挑战书～": {
        "图片": "mario-kart.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/attractions/mario-kart-koopas-challenge",
        "玩法": "戴上头箍和AR目镜后进入4人卡丁车，用方向盘上的按钮投掷龟壳、瞄准敌方角色并收集金币。车辆沿固定路线运行，不需要真正控制方向；视线追踪目标、在合适方向发射才是得分核心。",
        "操作诀窍": "上车前调紧头箍；看到敌人时先转头锁定再按按钮，连续乱按通常不如瞄准有效。两人同车更适合比较分数。",
        "官方参数": "约5分钟｜4名（2名×2排）｜单独122cm以上；有陪同者107cm以上｜AR、投影和空间演出",
        "实际占用": "速通约20—35分钟，包含佩戴设备和分组；普通队的库巴城布景更完整。",
        "提醒": "有急加速、旋转、急停、闪光和大音量；容易晕AR时可把视线更多放在实景。",
    },
    "耀西冒险": {
        "图片": "yoshi.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/attractions/yoshis-adventure",
        "玩法": "两人乘坐耀西车辆缓慢巡游，跟随奇诺比奥队长的地图寻找沿途3枚彩色蛋，并从高处观看蘑菇王国。节奏温和，价值主要在场景、拍照和找彩蛋。",
        "操作诀窍": "提前把手机调到快速拍照模式，入座后留意左右两侧和高处的蛋；排队超过55—60分钟时，通常不如把时间留给其他项目。",
        "官方参数": "约5分钟｜2名（2名×1排）｜单独122cm以上；有陪同者92cm以上｜亲子向",
        "实际占用": "普通队约45—90分钟；乘坐和上下车约10—15分钟。",
        "提醒": "任天堂区域可能需要入场确约券、整理券或抽选券；速通票面含区域入场时无需另抢同一时段。",
    },
    "哈利·波特与禁忌之旅": {
        "图片": "harry-potter.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/attractions/harry-potter-and-the-forbidden-journey/index.html",
        "玩法": "在霍格沃茨城堡内登上4人机械臂座椅，实景装置和巨幕影像交替出现，模拟追逐金色飞贼、魁地奇飞行、暴打柳和摄魂怪袭击。",
        "操作诀窍": "视线看向故事中心，头部靠稳，容易晕动时不要反复看画面边缘；想看完整城堡排队布景，可在速通体验后再走城堡漫步。",
        "官方参数": "约5分钟｜4名一排｜122cm以上｜机械臂、黑暗、影像、旋转和急停",
        "实际占用": "速通含存包通常约20—35分钟；单人通道是未买到该项目速通时的优先替代。",
        "提醒": "视觉运动感很强，眩晕、颈背不适者谨慎；乘坐前需使用项目免费寄存柜。",
    },
    "鹰马的飞行": {
        "图片": "hippogriff.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/attractions/flight-of-the-hippogriff",
        "玩法": "先经过海格小屋和南瓜地，向鹰马鞠躬后登上16人家庭过山车，在魔法世界上方完成一段短程飞行训练。",
        "操作诀窍": "路线很短，重点看海格小屋和城堡景观；App等待超过45分钟时跳过，除非你是哈迷或想补一个低强度过山车。",
        "官方参数": "约2分钟｜16名｜单独122cm以上且195cm以下；有陪同者92cm以上",
        "实际占用": "普通队常约45—90分钟；上下车后实际体验很短。",
        "提醒": "没有单人通道；虽是家庭项目，仍有下落、转弯和安全压杆。",
    },
    "霍格沃茨城堡漫步": {
        "图片": "harry-potter.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/events/the-wizarding-world-of-harry-potter/castle-walk",
        "玩法": "沿期间限定的步行路线进入霍格沃茨城堡，观看会说话的肖像、格兰芬多公共休息室入口和守护校长室的鹰兽雕像等细节，不乘坐禁忌之旅。",
        "操作诀窍": "紧接禁忌之旅安排，避免二次进出魔法世界；官方要求参观时持续向前，不要为了拍照长时间停留。",
        "官方参数": "2026年9月1日至2027年1月17日｜无身高限制｜轮椅可进入｜步行参观",
        "实际占用": "排队和参观通常约20—40分钟，以现场开放方式为准。",
        "提醒": "步行路线不包含禁忌之旅乘坐权；具体入口和开放时段需当天询问工作人员。",
    },
    "飞天翼龙": {
        "图片": "flying-dinosaur.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/attractions/the-flying-dinosaur",
        "玩法": "采用脸朝地面的俯卧悬挂姿势，被翼龙“抓住”后经历高速下落、翻转和360度旋转，重点是裸露感、失重和贴近地面的掠过感。",
        "操作诀窍": "所有随身物品先存柜，过金属探测；上车前取下容易掉落的饰品。情侣可走单人通道节省时间，出口汇合。",
        "官方参数": "约3分钟｜32名（4名×8排）｜132—198cm｜免费寄存柜＋金属探测",
        "实际占用": "单人约25—55分钟；普通队约60—110分钟；存包、安检和分组另需约15—20分钟。",
        "提醒": "强度很高；颈背、腰椎、血压、眩晕或脚踝仍有症状时不建议乘坐。",
    },
    "好莱坞美梦乘车游": {
        "图片": "hollywood-dream.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/attractions/hollywood-dream-the-ride",
        "玩法": "坐正向列车高速穿行好莱坞区域，每个座位可从当日曲目中选择BGM，让音乐和下落、转弯同步。",
        "操作诀窍": "入座后马上用面板选择曲目；开园先玩或使用单人通道，既能省时也不影响音乐体验。",
        "官方参数": "约3分钟｜36名（4名×9排）｜132cm以上｜免费寄存柜＋金属探测",
        "实际占用": "开园普通队约20—45分钟，午后可能70—120分钟；单人队视空座情况波动。",
        "提醒": "单人通道只适用于正向版本；曲目会调整，9月27日以当天列表为准。",
    },
    "好莱坞美梦～逆转世界～": {
        "图片": "hollywood-dream.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/attractions/hollywood-dream-the-ride",
        "玩法": "使用同一套轨道向后行驶，乘客看不到接下来的下落和转弯，未知感与漂浮感比正向更强。",
        "操作诀窍": "正向和逆向只选一个时，怕排队选正向、追求刺激选逆向；逆向没有单人通道，想玩应开园或买含它的速通。",
        "官方参数": "约3分钟｜132cm以上｜后向音乐过山车｜无单人通道",
        "实际占用": "普通队通常约80—140分钟；需另外计算存包和金属探测时间。",
        "提醒": "与正向共用部分设施但队列不同；看不到前方更容易产生眩晕。",
    },
    "侏罗纪公园乘船游～暗夜版～": {
        "图片": "jurassic-park.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/events/halloween-extreme-autumn-2026/jurassic-park-the-ride-in-the-dark",
        "玩法": "先乘船参观恐龙研究区，系统失控后进入黑暗设施躲避逃脱的恐龙，最终从25.9米高处俯冲落水；18:30后切换为低能见度暗夜版本。",
        "操作诀窍": "速通第四项若无固定时间，尽量18:30后使用；手机和易湿物品放防水袋，想少湿尽量不要坐最前排。",
        "官方参数": "常规乘船约7分钟｜最终落差25.9米｜9月11日至11月8日；9月27日暗夜版18:30后",
        "实际占用": "白天普通队约30—60分钟；夜间约50—90分钟；速通约15—30分钟。",
        "提醒": "是否能用所购速通体验暗夜版本，以9月27日票面和现场运行为准；一定可能湿身。",
    },
    "大白鲨～Red Alert～": {
        "图片": "jaws.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/attractions/jaws/",
        "玩法": "跟随船长参加阿米蒂港观光船之旅，途中遭遇巨型鲨鱼、枪击和火焰演出；Red Alert版本加入低照度、浓雾和红色水面。",
        "操作诀窍": "坐在船边更接近水面效果，中间位置相对不易湿；白天想了解剧情、晚上想要氛围，可根据排队二选一而非强求二刷。",
        "官方参数": "乘船约7分钟｜真人船长＋大型机械效果｜9月27日Red Alert于18:30后运行",
        "实际占用": "普通／单人通道约25—60分钟；夜间人潮集中时更长。",
        "提醒": "日语船长表演是体验核心，即使听不全也能靠动作理解；火焰靠近时温度明显。",
    },
    "小黄人疯狂乘车游": {
        "图片": "minion-mayhem.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/attractions/despicable-me-minion-mayhem/index.html",
        "玩法": "进入格鲁的住宅和实验室，搭乘8人动感座椅接受“变成小黄人”的训练，巨大穹幕影像与座椅运动共同制造加速、坠落和翻滚错觉。",
        "操作诀窍": "容易晕动就坐中间并看画面中心；项目总时长包含多个前置影片，不适合只想快速刷刺激项目的人。",
        "官方参数": "约25分钟｜8名（4名×2排）｜单独122cm以上；有陪同者102cm以上",
        "实际占用": "排队约40—80分钟，再加约25分钟体验；单人通道开放时可明显缩短。",
        "提醒": "视觉晃动强于实际移动，晕3D或模拟器者可以跳过。",
    },
    "小黄人大恶党任务": {
        "图片": "minion-mission.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/attractions/illuminations-villain-con-minion-blast",
        "玩法": "站在移动步道上使用E-Liminator X发射器，经过多个挑战房间射击、破坏或“偷走”目标获得分数；官方App可同步发射器、角色和成绩。",
        "操作诀窍": "不同目标使用对应按钮，不要只盯一个屏幕；入园前打开App中的Universal Play，NFC手机可同步设备并记录分数。",
        "官方参数": "约7分钟｜移动步道互动射击｜单独122cm以上；有陪同者无身高下限",
        "实际占用": "上午约25—50分钟；午后约40—70分钟，以App为准。",
        "提醒": "需要持续站立并随移动步道前进；脚踝容易疲劳时，把它放在早上而不是夜间。",
    },
    "名侦探柯南4-D现场秀": {
        "图片": "conan-4d.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/attractions/detective-conan-4d-live-show",
        "玩法": "在大型剧场观看柯南原创案件，3D影像与真人演员同台，并配合座椅运动、水、雾、闪光和大音量特殊效果。",
        "操作诀窍": "提前15—25分钟到场即可，靠中间位置更容易兼顾银幕和真人表演；日语理解有限会损失推理剧情，但动作场面仍直观。",
        "官方参数": "约30分钟｜750名｜3D＋真人演出｜无身高限制但有健康条件限制",
        "实际占用": "含候场约45—60分钟；适合作为午后坐下休息项目。",
        "提醒": "有闪光、雾、水、大音量、暗场及座椅运动；容易晕动或对闪光敏感者谨慎。",
    },
    "水世界": {
        "图片": "waterworld.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/attractions/waterworld",
        "玩法": "在大型水上基地观看摩托艇追逐、枪战、火焰、13米高空跳水和飞机冲入场地的真人特技秀。",
        "操作诀窍": "按App场次提前25—30分钟到场；前排蓝色湿身区水量很大，想休息和保持干燥就坐中后排。",
        "官方参数": "约20分钟｜3220名（含站席及轮椅区域）｜无身高限制｜大型爆破和水花",
        "实际占用": "候场加演出约45—55分钟；大容量使其比普通热门项目更容易规划。",
        "提醒": "9月27日具体场次尚未发布，路线中的14:30只是目标场次，需前一晚替换。",
    },
    "SING ON TOUR": {
        "图片": "sing-on-tour.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/attractions/sing-on-tour",
        "玩法": "在室内剧场观看Buster Moon主持的真人音乐剧，电影角色演唱熟悉曲目，并穿插舞台故障和剧情互动。",
        "操作诀窍": "不依赖日语也能理解音乐和表演，适合安排在17点左右恢复体力；提前15—20分钟进场即可。",
        "官方参数": "约20分钟｜506名（含轮椅区域）｜无身高限制｜室内坐席",
        "实际占用": "含候场约35—45分钟；比排一条中等队伍更适合作为休整。",
        "提醒": "场次会变，最终以9月27日官方App为准。",
    },
}

for attraction in USJ_GUIDE["attractions"]:
    attraction.update(USJ_ATTRACTION_DETAILS.get(attraction["名称"], {}))

USJ_HALLOWEEN_MEDIA: dict[str, dict[str, str]] = {
    "街头僵尸＋Zombie de Dance": {
        "图片": "halloween-night.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/events/halloween-extreme-autumn-2026/halloween-horror-nights",
    },
    "Factory of Fear 僵尸工厂": {
        "图片": "factory-of-fear.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/events/halloween-extreme-autumn-2026/halloween-horror-nights/factory-of-fear",
    },
    "生化危机 Requiem：The Dive": {
        "图片": "biohazard.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/events/halloween-extreme-autumn-2026/halloween-horror-nights/resident-evil-requiem-the-dive",
    },
    "电锯人 The Chaos 4-D": {
        "图片": "chainsaw-man.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/events/halloween-extreme-autumn-2026/halloween-horror-nights/chainsaw-man-the-chaos-4d",
    },
    "贞子的诅咒：暗黑恐怖乘车": {
        "图片": "sadako.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/events/halloween-extreme-autumn-2026/halloween-horror-nights/sadako",
    },
    "18号宅邸的魔女": {
        "图片": "witch-house.jpg",
        "官网": "https://www.usj.co.jp/web/ja/jp/events/halloween-extreme-autumn-2026/halloween-horror-nights/kate",
    },
}

for halloween_item in USJ_GUIDE["halloween"]:
    halloween_item.update(USJ_HALLOWEEN_MEDIA.get(halloween_item["名称"], {}))


def read_sheet(wb, name: str) -> list[dict[str, str]]:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(value or "") for value in rows[0]]
    result: list[dict[str, str]] = []
    for row in rows[1:]:
        item = {}
        for header, value in zip(headers, row):
            item[header] = "" if value is None else str(value)
        if any(item.values()):
            result.append(item)
    return result


def build_data() -> dict[str, object]:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    overview = read_sheet(wb, "12天总览")
    daily_food = read_sheet(wb, "每日吃法")
    food_map = read_sheet(wb, "美食推荐")
    lodging = read_sheet(wb, "住宿区域推荐")
    alternatives = read_sheet(wb, "P人备选方案")
    transport = read_sheet(wb, "交通预约")
    checklist = read_sheet(wb, "行前清单")
    daily_detail = read_sheet(wb, "关西加强每日攻略")

    for item in food_map:
        key = (item.get("区域", ""), item.get("类型", ""))
        item["地图店铺"] = FOOD_LOCATIONS.get(key, [])
    food_map.extend(EXTRA_FOOD_RECOMMENDATIONS)
    for item in lodging:
        item["住宿推荐"] = HOTEL_RECOMMENDATIONS.get(item.get("阶段", ""), [])

    foods_by_day = {item.get("日期", ""): item for item in daily_food}
    details_by_date: dict[str, list[dict[str, str]]] = {}
    for item in daily_detail:
        details_by_date.setdefault(item.get("日期", ""), []).append(item)

    days = []
    for item in overview:
        date = item.get("日期", "")
        short_date = date[5:] if len(date) >= 10 else date
        days.append(
            {
                **item,
                "短日期": short_date,
                "吃法": foods_by_day.get(short_date, {}),
                "细节": details_by_date.get(short_date, []),
                "交通分段": SEGMENT_ROUTES.get(item.get("天数", ""), []),
            }
        )

    return {
        "trip": {
            "title": "日本12天11晚",
            "subtitle": "广州出发 · 关西加强 · 东京收尾",
            "dates": "2026.09.25 - 2026.10.06",
            "route": "大阪4晚 → 京都4晚 → 东京3晚",
        },
        "days": days,
        "foodMap": food_map,
        "lodging": lodging,
        "alternatives": alternatives,
        "transport": transport,
        "checklist": checklist,
        "usj": USJ_GUIDE,
        "updatedAt": "2026-08-06",
    }


def write_site(data: dict[str, object]) -> None:
    ROOT.mkdir(exist_ok=True)
    (ROOT / "data.js").write_text(
        "window.TRIP_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    (ROOT / "index.html").write_text(
        """<!doctype html>
<html lang="zh-CN">
  <head>
    <meta charset="utf-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover" />
    <meta name="theme-color" content="#1e4d4a" />
    <title>日本12天11晚旅行攻略</title>
    <link rel="stylesheet" href="./styles.css" />
  </head>
  <body>
    <header class="hero" id="hero">
      <div class="hero__shade"></div>
      <nav class="topbar">
        <span class="brand">日本旅行</span>
        <a class="ghost-link" href="#today">行程</a>
      </nav>
      <section class="hero__content">
        <p class="eyebrow" id="tripDates"></p>
        <h1 id="tripTitle"></h1>
        <p class="subtitle" id="tripSubtitle"></p>
        <div class="route-pill" id="tripRoute"></div>
      </section>
    </header>

    <main>
      <section class="quick-band" aria-label="旅行摘要">
        <div class="quick-stat">
          <strong>12</strong>
          <span>天</span>
        </div>
        <div class="quick-stat">
          <strong>8</strong>
          <span>晚关西</span>
        </div>
        <div class="quick-stat">
          <strong>3</strong>
          <span>晚东京</span>
        </div>
      </section>

      <div class="tabs" role="tablist" aria-label="攻略分类">
        <button class="tab is-active" data-view="days" type="button">每日</button>
        <button class="tab" data-view="food" type="button">美食</button>
        <button class="tab" data-view="lodging" type="button">住宿</button>
        <button class="tab" data-view="usj" type="button">USJ攻略</button>
        <button class="tab" data-view="planb" type="button">备选</button>
        <button class="tab" data-view="prep" type="button">准备</button>
      </div>

      <section class="toolbar" id="dayToolbar">
        <label class="search">
          <span>搜索</span>
          <input id="searchInput" type="search" placeholder="京都、USJ、抹茶、富士山" />
        </label>
        <div class="chips" id="cityChips"></div>
      </section>

      <section class="content-view is-visible" id="daysView">
        <div class="section-head" id="today">
          <h2>每日行程</h2>
          <p>按当天体力调整，重点行程都留了删减口。</p>
        </div>
        <div class="day-list" id="dayList"></div>
      </section>

      <section class="content-view" id="foodView">
        <div class="section-head">
          <h2>美食推荐</h2>
          <p>按区域找吃的，点击店名可直接进入 Google Maps；排队太长就用备选。</p>
        </div>
        <div class="food-groups" id="foodGroups"></div>
      </section>

      <section class="content-view" id="lodgingView">
        <div class="section-head">
          <h2>住宿推荐</h2>
          <p>按人民币每晚约300–500元筛选；每家都能查看定位，或从你当前所在位置直接规划路线。</p>
        </div>
        <div class="simple-list" id="lodgingList"></div>
      </section>

      <section class="content-view" id="usjView">
        <div class="section-head">
          <h2>USJ项目与排队攻略</h2>
          <p>为9月28/29日整理：官方项目与活动规则＋2026历史数据估算。临出发前点击底部官方入口复核。</p>
        </div>
        <div class="usj-guide" id="usjGuide"></div>
      </section>

      <section class="content-view" id="planbView">
        <div class="section-head">
          <h2>备选方案</h2>
          <p>晚点、下雨、人多、富士山看不见时用。</p>
        </div>
        <div class="simple-list" id="planbList"></div>
      </section>

      <section class="content-view" id="prepView">
        <div class="section-head">
          <h2>准备清单</h2>
          <p>预约、交通、证件和行李放在这里。</p>
        </div>
        <div class="simple-list" id="transportList"></div>
        <div class="simple-list" id="checkList"></div>
      </section>
    </main>

    <footer>
      <span id="updatedAt"></span>
      <span>行程来自本地 Excel · 交通、美食与住宿定位核对 Google Maps</span>
    </footer>

    <script src="./data.js"></script>
    <script src="./app.js"></script>
  </body>
</html>
""",
        encoding="utf-8",
    )
    (ROOT / "styles.css").write_text(
        r""":root {
  color-scheme: light;
  --ink: #172321;
  --muted: #5f6f6b;
  --paper: #fffdf8;
  --band: #f2eee5;
  --line: #ded7c8;
  --green: #1e4d4a;
  --green-2: #2f6f65;
  --coral: #c85e40;
  --gold: #b3832d;
  --shadow: 0 14px 32px rgba(30, 44, 42, 0.12);
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
}

* {
  box-sizing: border-box;
}

html {
  scroll-behavior: smooth;
}

body {
  margin: 0;
  background: var(--paper);
  color: var(--ink);
}

.hero {
  min-height: 88svh;
  position: relative;
  display: flex;
  flex-direction: column;
  justify-content: space-between;
  background:
    radial-gradient(circle at 18% 16%, rgba(255, 255, 255, 0.22), transparent 25%),
    radial-gradient(circle at 88% 22%, rgba(200, 94, 64, 0.35), transparent 26%),
    linear-gradient(135deg, #173b39 0%, #1e4d4a 42%, #b3832d 100%);
  isolation: isolate;
}

.hero__shade {
  position: absolute;
  inset: 0;
  background:
    linear-gradient(180deg, rgba(0, 0, 0, 0.1), rgba(0, 0, 0, 0.04) 38%, rgba(0, 0, 0, 0.38) 100%),
    repeating-linear-gradient(120deg, rgba(255, 255, 255, 0.08) 0 1px, transparent 1px 18px);
  z-index: -1;
}

.topbar {
  display: flex;
  align-items: center;
  justify-content: space-between;
  padding: max(18px, env(safe-area-inset-top)) 18px 0;
  color: #fff;
}

.brand {
  font-size: 14px;
  font-weight: 700;
}

.ghost-link {
  color: #fff;
  text-decoration: none;
  border: 1px solid rgba(255, 255, 255, 0.62);
  border-radius: 999px;
  padding: 8px 14px;
  font-size: 13px;
  background: rgba(0, 0, 0, 0.18);
  backdrop-filter: blur(10px);
}

.hero__content {
  color: #fff;
  padding: 0 20px 28px;
  max-width: 780px;
}

.eyebrow {
  margin: 0 0 10px;
  font-size: 13px;
  font-weight: 700;
}

h1 {
  margin: 0;
  font-size: clamp(42px, 13vw, 86px);
  line-height: 0.94;
  letter-spacing: 0;
  overflow-wrap: anywhere;
}

.subtitle {
  max-width: 28rem;
  margin: 16px 0 18px;
  font-size: 17px;
  line-height: 1.55;
}

.route-pill {
  display: inline-flex;
  max-width: 100%;
  padding: 10px 14px;
  border-radius: 999px;
  background: rgba(255, 255, 255, 0.88);
  color: var(--green);
  font-size: 14px;
  font-weight: 800;
}

main {
  padding-bottom: 42px;
}

.quick-band {
  display: grid;
  grid-template-columns: repeat(3, 1fr);
  gap: 1px;
  background: var(--line);
  border-bottom: 1px solid var(--line);
}

.quick-stat {
  background: #faf6ed;
  padding: 16px 10px;
  text-align: center;
}

.quick-stat strong {
  display: block;
  font-size: 25px;
  color: var(--coral);
}

.quick-stat span {
  display: block;
  margin-top: 2px;
  font-size: 12px;
  color: var(--muted);
}

.tabs {
  position: sticky;
  top: 0;
  z-index: 8;
  display: flex;
  gap: 8px;
  overflow-x: auto;
  padding: 12px 14px;
  background: rgba(255, 253, 248, 0.94);
  border-bottom: 1px solid var(--line);
  backdrop-filter: blur(18px);
}

.tab {
  flex: 0 0 auto;
  border: 1px solid var(--line);
  background: #fff;
  color: var(--green);
  border-radius: 999px;
  padding: 9px 15px;
  font-weight: 800;
}

.tab.is-active {
  color: #fff;
  background: var(--green);
  border-color: var(--green);
}

.toolbar {
  padding: 14px 16px 4px;
}

.search {
  display: grid;
  gap: 6px;
  color: var(--muted);
  font-size: 12px;
  font-weight: 800;
}

.search input {
  width: 100%;
  min-height: 42px;
  border: 1px solid var(--line);
  border-radius: 8px;
  padding: 10px 12px;
  background: #fff;
  color: var(--ink);
  font: inherit;
}

.chips {
  display: flex;
  gap: 8px;
  overflow-x: auto;
  padding: 12px 0 4px;
}

.chip {
  flex: 0 0 auto;
  border: 1px solid var(--line);
  background: #fff;
  color: var(--muted);
  border-radius: 999px;
  padding: 8px 12px;
  font-size: 13px;
  font-weight: 700;
}

.chip.is-active {
  color: #fff;
  background: var(--coral);
  border-color: var(--coral);
}

.content-view {
  display: none;
}

.content-view.is-visible {
  display: block;
}

.section-head {
  padding: 22px 18px 10px;
}

.section-head h2 {
  margin: 0;
  font-size: 24px;
}

.section-head p {
  margin: 7px 0 0;
  color: var(--muted);
  line-height: 1.45;
}

.day-list,
.simple-list,
.food-groups {
  display: grid;
  gap: 12px;
  padding: 0 14px 18px;
}

.day-card,
.info-card,
.food-card {
  background: #fff;
  border: 1px solid var(--line);
  border-radius: 8px;
  box-shadow: var(--shadow);
  overflow: hidden;
}

.day-card summary {
  list-style: none;
  cursor: pointer;
  padding: 16px;
}

.day-card summary::-webkit-details-marker {
  display: none;
}

.day-top {
  display: flex;
  justify-content: space-between;
  gap: 12px;
  align-items: flex-start;
}

.day-kicker,
.tag {
  color: var(--coral);
  font-size: 12px;
  font-weight: 900;
}

.day-title {
  margin: 6px 0 0;
  font-size: 20px;
  line-height: 1.25;
}

.stay {
  color: var(--green);
  background: #e7f0ed;
  border-radius: 999px;
  padding: 7px 10px;
  font-size: 12px;
  font-weight: 800;
  white-space: nowrap;
}

.anchors {
  margin: 12px 0 0;
  color: var(--muted);
  line-height: 1.5;
}

.day-body {
  border-top: 1px solid var(--line);
  padding: 0 16px 16px;
}

.field {
  margin-top: 14px;
}

.field strong {
  display: block;
  margin-bottom: 4px;
  color: var(--green);
  font-size: 13px;
}

.field p {
  margin: 0;
  color: var(--ink);
  line-height: 1.55;
}

.detail-list {
  display: grid;
  gap: 8px;
  margin-top: 10px;
}

.route-list {
  display: grid;
  gap: 8px;
  margin-top: 8px;
}

.route-row {
  display: grid;
  grid-template-columns: minmax(0, 1fr) auto;
  gap: 4px 10px;
  align-items: center;
  padding: 10px 11px;
  border: 1px solid #dce8e4;
  border-radius: 7px;
  background: #f5faf8;
  color: var(--ink);
  text-decoration: none;
}

.route-row:active {
  background: #e7f0ed;
}

.route-option {
  display: inline-block;
  margin-right: 6px;
  color: var(--coral);
  font-size: 11px;
  font-weight: 900;
}

.route-points {
  min-width: 0;
  font-size: 14px;
  font-weight: 800;
  line-height: 1.4;
}

.route-time {
  color: var(--green);
  font-size: 13px;
  font-weight: 900;
  white-space: nowrap;
}

.route-note {
  grid-column: 1 / -1;
  color: var(--muted);
  font-size: 12px;
  line-height: 1.4;
}

.route-source {
  margin-top: 8px;
  color: var(--muted);
  font-size: 11px;
  line-height: 1.45;
}

.detail-row {
  border-left: 3px solid var(--gold);
  padding: 4px 0 4px 10px;
}

.info-card,
.food-card {
  padding: 15px;
}

.info-card h3,
.food-card h3 {
  margin: 0 0 8px;
  font-size: 17px;
  line-height: 1.35;
}

.meta {
  display: flex;
  flex-wrap: wrap;
  gap: 6px;
  margin-bottom: 10px;
}

.meta span {
  border-radius: 999px;
  padding: 5px 8px;
  background: #f4efe4;
  color: var(--muted);
  font-size: 12px;
  font-weight: 700;
}

.food-region {
  padding: 14px 4px 0;
  color: var(--green);
  font-size: 18px;
}

.food-map-links {
  display: flex;
  flex-wrap: wrap;
  gap: 7px;
  margin: 10px 0 2px;
}

.food-map-link {
  display: inline-flex;
  align-items: center;
  gap: 4px;
  max-width: 100%;
  padding: 7px 9px;
  border: 1px solid #cddfda;
  border-radius: 999px;
  background: #edf6f3;
  color: var(--green);
  font-size: 12px;
  font-weight: 850;
  line-height: 1.3;
  text-decoration: none;
}

.food-map-link:active {
  background: #dcebe7;
}

.food-map-note {
  width: 100%;
  color: var(--muted);
  font-size: 11px;
  line-height: 1.4;
}

.hotel-list {
  display: grid;
  gap: 9px;
  margin-top: 9px;
}

.hotel-pick {
  padding: 11px;
  border: 1px solid #dce8e4;
  border-radius: 7px;
  background: #f7faf9;
}

.hotel-top {
  display: flex;
  justify-content: space-between;
  gap: 8px;
  align-items: flex-start;
}

.hotel-name {
  margin: 0;
  font-size: 14px;
  line-height: 1.4;
}

.hotel-budget {
  flex: 0 0 auto;
  padding: 4px 7px;
  border-radius: 999px;
  background: #f4efe4;
  color: var(--muted);
  font-size: 11px;
  font-weight: 800;
}

.hotel-area {
  margin-top: 3px;
  color: var(--coral);
  font-size: 11px;
  font-weight: 900;
}

.hotel-reason {
  margin: 7px 0 0;
  color: var(--ink);
  font-size: 12px;
  line-height: 1.5;
}

.hotel-actions {
  display: flex;
  flex-wrap: wrap;
  gap: 7px;
  margin-top: 9px;
}

.hotel-link {
  display: inline-flex;
  align-items: center;
  padding: 7px 9px;
  border: 1px solid #cddfda;
  border-radius: 999px;
  background: #fff;
  color: var(--green);
  font-size: 12px;
  font-weight: 850;
  text-decoration: none;
}

.hotel-link--route {
  border-color: var(--green);
  background: var(--green);
  color: #fff;
}

.hotel-location-note {
  margin-top: 9px;
  color: var(--muted);
  font-size: 11px;
  line-height: 1.45;
}

.usj-guide {
  display: grid;
  gap: 14px;
  padding: 0 14px 22px;
}

.usj-hero-card,
.usj-section,
.usj-date-card,
.usj-strategy-card,
.usj-pass-card,
.usj-single-card {
  border: 1px solid var(--line);
  border-radius: 8px;
  background: #fff;
  box-shadow: var(--shadow);
}

.usj-hero-card {
  padding: 16px;
  color: #fff;
  background: linear-gradient(135deg, #c85e40 0%, #8f2f34 50%, #3b1e32 100%);
  border: 0;
}

.usj-hero-card h3,
.usj-section h3 {
  margin: 0;
}

.usj-hero-card p {
  margin: 8px 0 0;
  line-height: 1.55;
}

.usj-hero-card .usj-small-note {
  color: rgba(255, 255, 255, 0.82);
}

.usj-section {
  padding: 15px;
}

.usj-section-head {
  display: flex;
  justify-content: space-between;
  gap: 12px;
  align-items: baseline;
  margin-bottom: 11px;
}

.usj-section-head span,
.usj-small-note {
  color: var(--muted);
  font-size: 11px;
  line-height: 1.45;
}

.usj-date-grid,
.usj-strategy-grid,
.usj-pass-grid,
.usj-single-grid {
  display: grid;
  gap: 9px;
}

.usj-date-card,
.usj-strategy-card,
.usj-pass-card,
.usj-single-card {
  padding: 12px;
  box-shadow: none;
  background: #f9fbfa;
}

.usj-date-card.is-primary {
  border-color: var(--coral);
  background: #fff8f5;
}

.usj-date-top,
.usj-attraction-top {
  display: flex;
  justify-content: space-between;
  gap: 10px;
  align-items: flex-start;
}

.usj-date-top h4,
.usj-strategy-card h4,
.usj-pass-card h4,
.usj-single-card h4 {
  margin: 0;
  font-size: 15px;
}

.usj-status,
.usj-wait,
.usj-decision,
.usj-single-status {
  flex: 0 0 auto;
  border-radius: 999px;
  padding: 5px 8px;
  font-size: 11px;
  font-weight: 900;
}

.usj-status,
.usj-decision {
  color: #fff;
  background: var(--coral);
}

.usj-wait {
  color: var(--green);
  background: #e7f0ed;
  white-space: nowrap;
}

.usj-date-meta,
.usj-strategy-meta {
  display: grid;
  gap: 5px;
  margin-top: 9px;
  color: var(--muted);
  font-size: 12px;
  line-height: 1.45;
}

.usj-attraction-list,
.usj-halloween-list,
.usj-source-list {
  display: grid;
  gap: 8px;
}

.usj-attraction {
  border: 1px solid #dce8e4;
  border-radius: 8px;
  overflow: hidden;
  background: #fff;
}

.usj-attraction-image {
  display: block;
  width: 100%;
  aspect-ratio: 7 / 3;
  object-fit: cover;
  background: #e7efec;
  border-bottom: 1px solid var(--line);
}

.usj-attraction summary {
  list-style: none;
  cursor: pointer;
  padding: 12px;
}

.usj-attraction summary::-webkit-details-marker {
  display: none;
}

.usj-attraction-title {
  min-width: 0;
}

.usj-attraction-title h4 {
  margin: 0;
  font-size: 15px;
  line-height: 1.35;
}

.usj-attraction-title p {
  margin: 3px 0 0;
  color: var(--muted);
  font-size: 11px;
}

.usj-rating {
  margin-top: 8px;
  color: var(--gold);
  font-size: 13px;
  letter-spacing: 1px;
}

.usj-attraction-body {
  padding: 0 12px 12px;
  border-top: 1px solid var(--line);
}

.usj-official-link,
.usj-official-map-link {
  display: inline-flex;
  align-items: center;
  justify-content: center;
  min-height: 38px;
  margin-top: 10px;
  padding: 8px 12px;
  border: 1px solid var(--green);
  border-radius: 999px;
  color: var(--green);
  background: #fff;
  font-size: 12px;
  font-weight: 850;
  text-decoration: none;
}

.usj-halloween-item {
  display: grid;
  grid-template-columns: minmax(112px, 34%) 1fr;
  gap: 11px;
  padding: 0;
  border-left: 4px solid #8f2f34;
  border-radius: 0 7px 7px 0;
  background: #fbf5f6;
  overflow: hidden;
}

.usj-halloween-item > img {
  width: 100%;
  height: 100%;
  min-height: 128px;
  object-fit: cover;
}

.usj-halloween-copy {
  padding: 11px 11px 11px 0;
}

.usj-halloween-item h4 {
  margin: 0;
  font-size: 14px;
}

.usj-halloween-item p {
  margin: 5px 0 0;
  font-size: 12px;
  line-height: 1.5;
}

.usj-halloween-item a {
  display: inline-block;
  margin-top: 7px;
  color: #8f2f34;
  font-size: 11px;
  font-weight: 850;
  text-decoration: none;
}

.usj-map-scroll {
  overflow-x: auto;
  overscroll-behavior-inline: contain;
  padding: 4px 0 8px;
  scrollbar-width: thin;
}

.usj-route-map {
  display: block;
  width: 100%;
  min-width: 760px;
  height: auto;
}

.usj-map-zone rect {
  fill: rgba(255, 255, 255, 0.9);
  stroke: #8db4a8;
  stroke-width: 3;
}

.usj-map-zone text {
  fill: #244b42;
  text-anchor: middle;
  font-size: 16px;
  font-weight: 800;
}

.usj-map-zone text + text {
  fill: #607a73;
  font-size: 12px;
  font-weight: 650;
}

.usj-map-zone--nintendo rect {
  fill: #fff3c6;
  stroke: #d59b22;
}

.usj-map-zone--harry rect {
  fill: #eae6f5;
  stroke: #796c9e;
}

.usj-map-zone--minion rect {
  fill: #fff7b7;
  stroke: #d4a900;
}

.usj-map-zone--hollywood rect {
  fill: #fbe7e1;
  stroke: #d26b51;
}

.usj-map-entrance path {
  fill: #274f46;
}

.usj-map-entrance text {
  fill: #274f46;
  text-anchor: middle;
  font-size: 13px;
  font-weight: 900;
}

.usj-map-route-line {
  fill: none;
  stroke: #d65f42;
  stroke-width: 8;
  stroke-linecap: round;
  stroke-linejoin: round;
  stroke-dasharray: 3 15;
  opacity: 0.92;
}

.usj-map-marker circle {
  fill: #d65f42;
  stroke: #fff;
  stroke-width: 4;
}

.usj-map-marker text {
  fill: #fff;
  text-anchor: middle;
  font-size: 13px;
  font-weight: 950;
}

.usj-map-legend {
  display: flex;
  flex-wrap: wrap;
  gap: 8px 14px;
  margin-top: 7px;
  color: var(--muted);
  font-size: 11px;
}

.usj-map-legend span {
  display: inline-flex;
  align-items: center;
  gap: 6px;
}

.usj-legend-line {
  width: 24px;
  border-top: 3px dashed #d65f42;
}

.usj-legend-dot {
  width: 12px;
  height: 12px;
  border-radius: 50%;
  background: #d65f42;
  box-shadow: 0 0 0 2px #fff, 0 0 0 3px #d65f42;
}

.usj-official-map-link {
  width: 100%;
  color: #fff;
  background: var(--green);
}

.usj-route-assumption {
  padding: 12px;
  border: 1px solid #e7c981;
  border-radius: 8px;
  background: #fff8e9;
}

.usj-route-assumption > strong {
  display: block;
  color: #7a4f13;
  font-size: 14px;
}

.usj-route-assumption p {
  margin: 6px 0 0;
  color: #765b24;
  font-size: 11px;
  line-height: 1.5;
}

.usj-route-timeline {
  position: relative;
  display: grid;
  gap: 10px;
  margin-top: 13px;
}

.usj-route-timeline::before {
  content: "";
  position: absolute;
  top: 18px;
  bottom: 18px;
  left: 17px;
  width: 3px;
  background: linear-gradient(#d65f42, #2e675a);
}

.usj-route-step {
  position: relative;
  display: grid;
  grid-template-columns: 36px minmax(0, 1fr);
  gap: 10px;
}

.usj-route-step-marker {
  z-index: 1;
  display: grid;
  place-items: center;
  width: 36px;
  height: 36px;
  border: 3px solid #fff;
  border-radius: 50%;
  color: #fff;
  background: #d65f42;
  box-shadow: 0 0 0 1px #d65f42;
  font-size: 13px;
  font-weight: 950;
}

.usj-route-step-content {
  min-width: 0;
  padding: 11px;
  border: 1px solid #dce8e4;
  border-radius: 8px;
  background: #fbfdfc;
}

.usj-route-step-head {
  display: flex;
  justify-content: space-between;
  gap: 10px;
  align-items: flex-start;
}

.usj-route-step-head span {
  color: #c1553c;
  font-size: 11px;
  font-weight: 900;
}

.usj-route-step-head h4 {
  margin: 3px 0 0;
  font-size: 14px;
  line-height: 1.4;
}

.usj-route-step-head em {
  flex: 0 0 auto;
  padding: 4px 7px;
  border-radius: 999px;
  color: var(--green);
  background: #e7f0ed;
  font-size: 10px;
  font-style: normal;
  font-weight: 850;
}

.usj-route-facts {
  display: grid;
  gap: 4px;
  margin-top: 8px;
  padding: 8px;
  border-radius: 6px;
  background: #f0f6f4;
  color: var(--muted);
  font-size: 11px;
  line-height: 1.45;
}

.usj-strategy-card.is-recommended {
  border-color: var(--green);
  background: #f2f8f6;
}

.usj-pass-card.is-recommended {
  border-color: var(--coral);
  background: #fff8f5;
}

.usj-pass-projects {
  margin: 10px 0 0;
  padding-left: 21px;
  font-size: 12px;
  line-height: 1.55;
}

.usj-pass-projects li + li {
  margin-top: 3px;
}

.usj-single-card p {
  margin: 7px 0 0;
  color: var(--muted);
  font-size: 12px;
  line-height: 1.55;
}

.usj-single-status {
  color: var(--green);
  background: #e7f0ed;
  white-space: nowrap;
}

.usj-single-rules {
  margin-top: 12px;
  padding-top: 12px;
  border-top: 1px dashed var(--line);
}

.usj-inline-note {
  margin: 11px 0 0;
  padding: 9px 10px;
  border-radius: 7px;
  background: #fff8e9;
  color: #765b24;
  font-size: 11px;
  line-height: 1.5;
}

.usj-closure-list {
  margin: 0;
  padding-left: 19px;
  color: var(--ink);
  font-size: 12px;
  line-height: 1.55;
}

.usj-closure-list li + li {
  margin-top: 6px;
}

.usj-source-link {
  display: flex;
  justify-content: space-between;
  gap: 10px;
  padding: 10px 11px;
  border: 1px solid #cddfda;
  border-radius: 7px;
  background: #f5faf8;
  color: var(--green);
  font-size: 12px;
  font-weight: 850;
  text-decoration: none;
}

.usj-source-link::after {
  content: "打开 ↗";
  flex: 0 0 auto;
}

footer {
  display: flex;
  justify-content: space-between;
  gap: 12px;
  padding: 24px 18px calc(24px + env(safe-area-inset-bottom));
  background: var(--green);
  color: rgba(255, 255, 255, 0.82);
  font-size: 12px;
}

@media (min-width: 760px) {
  .hero {
    min-height: 72vh;
  }

  .hero__content {
    padding-left: 7vw;
    padding-bottom: 7vh;
  }

  main {
    max-width: 980px;
    margin: 0 auto;
  }

  .quick-band {
    border-left: 1px solid var(--line);
    border-right: 1px solid var(--line);
  }

  .day-list,
  .simple-list,
  .food-groups,
  .usj-date-grid,
  .usj-strategy-grid {
    grid-template-columns: repeat(2, minmax(0, 1fr));
  }

  .day-list {
    grid-template-columns: 1fr;
  }
}
""",
        encoding="utf-8",
    )
    (ROOT / "app.js").write_text(
        r"""const data = window.TRIP_DATA;

const state = {
  view: "days",
  city: "全部",
  query: "",
};

const $ = (selector) => document.querySelector(selector);

function text(value) {
  return value || "按当天情况调整";
}

function googleMapsUrl(route) {
  const params = new URLSearchParams({
    api: "1",
    origin: route["地图起点"],
    destination: route["地图终点"],
    travelmode: route["方式"] === "步行" ? "walking" : "transit",
  });
  return `https://www.google.com/maps/dir/?${params.toString()}`;
}

function routeList(routes) {
  return `
    <div class="route-list">
      ${routes.map(route => `
        <a class="route-row" href="${googleMapsUrl(route)}" target="_blank" rel="noopener noreferrer" title="在 Google Maps 打开实时路线">
          <div class="route-points">
            ${route["方案"] ? `<span class="route-option">${route["方案"]}</span>` : ""}
            ${route["起点"]} → ${route["终点"]}
          </div>
          <span class="route-time">${route["方式"]} · ${route["耗时"]}</span>
          ${route["说明"] ? `<span class="route-note">${route["说明"]}</span>` : ""}
        </a>
      `).join("")}
    </div>
    <div class="route-source">Google Maps 于 2026-07-30 日间查询；点击任一路段可按出发时刻查看实时班次。未含逛景点、排队及大型车站内找路时间。</div>
  `;
}

function includesQuery(item, query) {
  if (!query) return true;
  return JSON.stringify(item).toLowerCase().includes(query.toLowerCase());
}

function detectCity(day) {
  const blob = `${day["入住"]} ${day["主线"]} ${day["必去锚点"]}`;
  if (blob.includes("大阪") || blob.includes("USJ") || blob.includes("神户") || blob.includes("姬路")) return "关西";
  if (blob.includes("京都") || blob.includes("宇治") || blob.includes("奈良")) return "京都";
  if (blob.includes("东京") || blob.includes("富士") || blob.includes("镰仓")) return "东京";
  return "全部";
}

function renderHero() {
  $("#tripDates").textContent = data.trip.dates;
  $("#tripTitle").textContent = data.trip.title;
  $("#tripSubtitle").textContent = data.trip.subtitle;
  $("#tripRoute").textContent = data.trip.route;
  $("#updatedAt").textContent = `更新 ${data.updatedAt}`;
}

function renderChips() {
  const chips = ["全部", "关西", "京都", "东京"];
  $("#cityChips").innerHTML = chips
    .map((city) => `<button class="chip ${state.city === city ? "is-active" : ""}" data-city="${city}" type="button">${city}</button>`)
    .join("");
}

function dayCard(day, index) {
  const food = day["吃法"] || {};
  const details = day["细节"] || [];
  const routes = day["交通分段"] || [];
  const open = index === 0 ? "open" : "";
  return `
    <details class="day-card" ${open}>
      <summary>
        <div class="day-top">
          <div>
            <div class="day-kicker">${day["天数"]} · ${day["日期"]} · ${day["星期"]}</div>
            <h3 class="day-title">${day["主线"]}</h3>
          </div>
          <span class="stay">${day["入住"]}</span>
        </div>
        <p class="anchors">${day["必去锚点"]}</p>
      </summary>
      <div class="day-body">
        <div class="field"><strong>P人玩法</strong><p>${day["P人友好玩法"]}</p></div>
        <div class="field"><strong>交通</strong><p>${day["交通重点"]}</p></div>
        ${routes.length ? `<div class="field"><strong>逐段交通耗时 · ${routes.length}段</strong>${routeList(routes)}</div>` : ""}
        <div class="field"><strong>晚餐/夜间</strong><p>${day["晚餐/夜间建议"]}</p></div>
        <div class="field"><strong>当天吃法</strong><p>${text(food["午餐"])}；${text(food["晚餐"])}。${text(food["P人吃法"])}</p></div>
        ${details.length ? `<div class="field"><strong>分时段</strong><div class="detail-list">${details.map(detail => `
          <div class="detail-row">
            <div class="tag">${detail["时间块"]} · ${detail["城市"]}</div>
            <p>${detail["建议安排"]}</p>
          </div>`).join("")}</div></div>` : ""}
      </div>
    </details>
  `;
}

function renderDays() {
  const filtered = data.days.filter((day) => {
    const cityOk = state.city === "全部" || detectCity(day) === state.city;
    return cityOk && includesQuery(day, state.query);
  });
  $("#dayList").innerHTML = filtered.map(dayCard).join("") || `<div class="info-card"><h3>没有匹配内容</h3><p>换个关键词试试。</p></div>`;
}

function googleMapsSearchUrl(location) {
  const params = new URLSearchParams({
    api: "1",
    query: location["地图查询"],
  });
  return `https://www.google.com/maps/search/?${params.toString()}`;
}

function foodLocationLinks(locations) {
  if (!locations.length) return "";
  return `
    <div class="food-map-links">
      ${locations.map(location => `
        <a class="food-map-link" href="${googleMapsSearchUrl(location)}" target="_blank" rel="noopener noreferrer" title="在 Google Maps 打开 ${location["店名"]}">
          <span aria-hidden="true">📍</span>${location["店名"]}<span aria-hidden="true">↗</span>
        </a>
        ${location["定位说明"] ? `<span class="food-map-note">${location["定位说明"]}</span>` : ""}
      `).join("")}
    </div>
  `;
}

function renderFood() {
  const groups = data.foodMap.reduce((acc, item) => {
    const region = item["区域"] || "其他";
    acc[region] = acc[region] || [];
    acc[region].push(item);
    return acc;
  }, {});
  $("#foodGroups").innerHTML = Object.entries(groups).map(([region, items]) => `
    <div class="food-region">${region}</div>
    ${items.map(item => `
      <article class="food-card">
        <div class="meta"><span>${item["类型"]}</span><span>${item["预算感"]}</span><span>${item["适合日期"]}</span></div>
        <h3>${item["推荐店/吃法"]}</h3>
        ${foodLocationLinks(item["地图店铺"] || [])}
        <div class="field"><strong>预约/排队</strong><p>${item["预约/排队"]}</p></div>
        <div class="field"><strong>点单</strong><p>${item["点单建议"]}</p></div>
        <div class="field"><strong>备选</strong><p>${item["P人备选"]}</p></div>
      </article>
    `).join("")}
  `).join("");
}

function googleMapsFromHereUrl(hotel) {
  const params = new URLSearchParams({
    api: "1",
    destination: hotel["地图查询"],
    travelmode: "transit",
  });
  return `https://www.google.com/maps/dir/?${params.toString()}`;
}

function hotelList(hotels) {
  if (!hotels.length) return "";
  return `
    <div class="hotel-list">
      ${hotels.map(hotel => `
        <div class="hotel-pick">
          <div class="hotel-top">
            <div>
              <h4 class="hotel-name">${hotel["名称"]}</h4>
              <div class="hotel-area">${hotel["区域"]}</div>
            </div>
            <span class="hotel-budget">参考 ${hotel["预算感"]}</span>
          </div>
          <p class="hotel-reason">${hotel["适合理由"]}</p>
          <div class="hotel-actions">
            <a class="hotel-link" href="${googleMapsSearchUrl(hotel)}" target="_blank" rel="noopener noreferrer" title="在 Google Maps 查看 ${hotel["名称"]}">📍 地图定位</a>
            <a class="hotel-link hotel-link--route" href="${googleMapsFromHereUrl(hotel)}" target="_blank" rel="noopener noreferrer" title="从当前位置前往 ${hotel["名称"]}">从我这里出发 ↗</a>
          </div>
        </div>
      `).join("")}
    </div>
    <div class="hotel-location-note">“从我这里出发”不会写死起点；手机需允许 Google Maps 使用当前位置。价格按单人或一间基础房估算，仅作 ¥300–500 筛选参考；请以 2026年9月25日至10月6日 的实时含税总价为准。标有共用卫浴、青旅或胶囊的候选，下单前请再次确认房型。</div>
  `;
}

function lodgingCard(item) {
  const hotels = item["住宿推荐"] || [];
  return `
    <article class="info-card">
      <div class="meta">
        <span>${item["日期"]}</span>
        <span>${item["晚数"]}晚</span>
        <span>备选 ${item["备选区域"]}</span>
      </div>
      <h3>${item["阶段"]} · ${item["首选区域"]}</h3>
      <div class="field"><strong>为什么</strong><p>${item["为什么"]}</p></div>
      <div class="field"><strong>导游建议</strong><p>${item["导游建议"]}</p></div>
      ${hotels.length ? `<div class="field"><strong>具体住宿 · ${hotels.length}家</strong>${hotelList(hotels)}</div>` : ""}
    </article>
  `;
}

function simpleCard(title, meta, fields) {
  return `
    <article class="info-card">
      <div class="meta">${meta.filter(Boolean).map(item => `<span>${item}</span>`).join("")}</div>
      <h3>${title}</h3>
      ${fields.map(([label, value]) => `<div class="field"><strong>${label}</strong><p>${value}</p></div>`).join("")}
    </article>
  `;
}

function usjRating(score) {
  return "★".repeat(score) + "☆".repeat(5 - score);
}

function usjAssetPath(fileName) {
  const inTravelPage = /\/travel-page\/(?:index\.html)?$/.test(window.location.pathname);
  return inTravelPage ? `./assets/usj/${fileName}` : `./travel-page/assets/usj/${fileName}`;
}

function renderUSJRouteMap() {
  return `
    <div class="usj-map-scroll" role="region" aria-label="USJ园区路线示意图，可横向滚动">
      <svg class="usj-route-map" viewBox="0 0 900 660" role="img" aria-labelledby="usjMapTitle usjMapDesc">
        <title id="usjMapTitle">USJ园区逆时针游玩路线示意图</title>
        <desc id="usjMapDesc">从入口出发，依次经过好莱坞、小黄人、超级任天堂、侏罗纪、水世界、亲善村、哈利波特，再回到前区和万圣节夜间项目。</desc>
        <defs>
          <linearGradient id="usjMapBg" x1="0" y1="0" x2="1" y2="1">
            <stop offset="0" stop-color="#eef8f5" />
            <stop offset="1" stop-color="#fff6e6" />
          </linearGradient>
          <marker id="usjArrow" viewBox="0 0 10 10" refX="9" refY="5" markerWidth="7" markerHeight="7" orient="auto-start-reverse">
            <path d="M 0 0 L 10 5 L 0 10 z" fill="#d65f42" />
          </marker>
        </defs>
        <rect x="8" y="8" width="884" height="644" rx="30" fill="url(#usjMapBg)" stroke="#b8d3ca" stroke-width="4" />
        <path d="M30 80 C180 20 340 60 450 28 C600 -5 760 35 870 92 L870 330 C810 300 780 310 735 360 C650 455 555 585 450 630 C330 600 260 535 180 505 C95 475 45 390 30 300 Z" fill="#d8ece7" opacity="0.55" />

        <g class="usj-map-zone usj-map-zone--nintendo"><rect x="34" y="48" width="170" height="105" rx="22" /><text x="119" y="91">超级任天堂世界</text><text x="119" y="119">咚奇刚 · 马里奥 · 耀西</text></g>
        <g class="usj-map-zone"><rect x="68" y="205" width="190" height="118" rx="22" /><text x="163" y="249">侏罗纪公园</text><text x="163" y="277">飞天翼龙 · 乘船游</text></g>
        <g class="usj-map-zone"><rect x="280" y="45" width="165" height="98" rx="22" /><text x="362" y="88">水世界</text><text x="362" y="113">大型特技秀</text></g>
        <g class="usj-map-zone"><rect x="470" y="88" width="165" height="106" rx="22" /><text x="552" y="131">亲善村</text><text x="552" y="158">大白鲨</text></g>
        <g class="usj-map-zone usj-map-zone--harry"><rect x="650" y="148" width="205" height="132" rx="22" /><text x="752" y="197">哈利·波特</text><text x="752" y="225">禁忌之旅 · 鹰马</text><text x="752" y="250">城堡漫步</text></g>
        <g class="usj-map-zone"><rect x="658" y="332" width="180" height="96" rx="22" /><text x="748" y="373">奇境世界</text><text x="748" y="399">亲子项目区</text></g>
        <g class="usj-map-zone usj-map-zone--minion"><rect x="66" y="384" width="190" height="105" rx="22" /><text x="161" y="426">小黄人乐园</text><text x="161" y="454">大恶党任务 · 乘车游</text></g>
        <g class="usj-map-zone"><rect x="278" y="374" width="150" height="96" rx="22" /><text x="353" y="415">旧金山</text><text x="353" y="442">餐饮休整</text></g>
        <g class="usj-map-zone"><rect x="440" y="386" width="175" height="96" rx="22" /><text x="527" y="426">纽约</text><text x="527" y="453">柯南4-D</text></g>
        <g class="usj-map-zone usj-map-zone--hollywood"><rect x="318" y="500" width="260" height="105" rx="22" /><text x="448" y="542">好莱坞</text><text x="448" y="570">美梦乘车游 · SING</text></g>
        <g class="usj-map-entrance"><path d="M398 635 H502 L480 610 H420 Z" /><text x="450" y="649">入口／出口</text></g>

        <polyline class="usj-map-route-line" points="450,620 450,520 160,440 120,105 165,260 185,275 362,95 552,145 752,215 500,435 190,260" marker-end="url(#usjArrow)" />
        ${[
          [1, 450, 620], [2, 450, 520], [3, 160, 440], [4, 120, 105], [5, 165, 242],
          [6, 190, 285], [7, 362, 95], [8, 552, 145], [9, 752, 215], [10, 500, 435], [11, 190, 260]
        ].map(([number, x, y]) => `<g class="usj-map-marker"><circle cx="${x}" cy="${y}" r="17" /><text x="${x}" y="${y + 5}">${number}</text></g>`).join("")}
      </svg>
    </div>
    <div class="usj-map-legend">
      <span><i class="usj-legend-line"></i>推荐步行方向</span>
      <span><i class="usj-legend-dot"></i>数字对应下方时间轴</span>
    </div>
    <p class="usj-inline-note">这是为减少折返绘制的路线示意图，不是精确比例地图。现场定位、洗手间和临时封路请打开USJ官方App地图。</p>
  `;
}

function renderUSJ() {
  const guide = data.usj;
  if (!guide) return;

  const dates = guide.dates.map((item, index) => `
    <article class="usj-date-card ${index === 0 ? "is-primary" : ""}">
      <div class="usj-date-top">
        <h4>${item["日期"]}</h4>
        <span class="usj-status">${item["建议"]}</span>
      </div>
      <div class="usj-date-meta">
        <span>营业：${item["营业时间"]}</span>
        <span>预测：${item["预测平均等待"]} · ${item["年卡情况"]}</span>
        <span>${item["行程影响"]}</span>
      </div>
    </article>
  `).join("");

  const attractions = guide.attractions.map((item, index) => `
    <details class="usj-attraction" ${index < 2 ? "open" : ""}>
      <img class="usj-attraction-image" src="${usjAssetPath(item["图片"])}" alt="${item["名称"]}官方项目图片" loading="lazy" decoding="async" />
      <summary>
        <div class="usj-attraction-top">
          <div class="usj-attraction-title">
            <h4>${item["名称"]}</h4>
            <p>${item["区域"]} · ${item["类型"]}</p>
          </div>
          <span class="usj-wait">${item["预计排队"]}</span>
        </div>
        <div class="usj-rating" aria-label="推荐度${item["推荐度"]}星">${usjRating(item["推荐度"])}</div>
      </summary>
      <div class="usj-attraction-body">
        <div class="field"><strong>项目介绍</strong><p>${item["介绍"]}</p></div>
        <div class="field"><strong>怎么玩</strong><p>${item["玩法"]}</p></div>
        <div class="field"><strong>玩法诀窍</strong><p>${item["操作诀窍"]}</p></div>
        <div class="field"><strong>官网参数</strong><p>${item["官方参数"]}</p></div>
        <div class="field"><strong>预计实际占用</strong><p>${item["实际占用"]}</p></div>
        <div class="field"><strong>乘坐提醒</strong><p>${item["提醒"]}</p></div>
        <div class="field"><strong>速通建议</strong><p>${item["速通建议"]}</p></div>
        <div class="field"><strong>单人通道</strong><p>${item["单人通道"]}</p></div>
        <div class="field"><strong>最终取舍</strong><p>${item["取舍"]}</p></div>
        <a class="usj-official-link" href="${item["官网"]}" target="_blank" rel="noopener noreferrer">打开USJ官方项目页</a>
      </div>
    </details>
  `).join("");

  const halloween = guide.halloween.map(item => `
    <article class="usj-halloween-item">
      <img src="${usjAssetPath(item["图片"])}" alt="${item["名称"]}官方项目图片" loading="lazy" decoding="async" />
      <div class="usj-halloween-copy">
      <h4>${item["名称"]}</h4>
      <p><strong>${item["时间"]}</strong> · ${item["预计等待"]}</p>
      <p>${item["建议"]}</p>
      <a href="${item["官网"]}" target="_blank" rel="noopener noreferrer">查看官网详情 ↗</a>
      </div>
    </article>
  `).join("");

  const routeTimeline = guide.route_timeline.map(item => `
    <article class="usj-route-step">
      <div class="usj-route-step-marker">${item["序号"]}</div>
      <div class="usj-route-step-content">
        <div class="usj-route-step-head">
          <div><span>${item["时间"]}</span><h4>${item["安排"]}</h4></div>
          <em>${item["区域"]}</em>
        </div>
        <div class="usj-route-facts">
          <span><strong>排队：</strong>${item["预计等待"]}</span>
          <span><strong>占用：</strong>${item["游玩占用"]}</span>
          <span><strong>通道：</strong>${item["通道"]}</span>
        </div>
        <div class="field"><strong>为什么这样排</strong><p>${item["理由"]}</p></div>
        <div class="field"><strong>现场调整</strong><p>${item["调整"]}</p></div>
      </div>
    </article>
  `).join("");

  const strategies = guide.strategies.map(item => `
    <article class="usj-strategy-card ${item["方案"].includes("Pass 4") ? "is-recommended" : ""}">
      <div class="usj-date-top">
        <h4>${item["方案"]}</h4>
        ${item["方案"].includes("Pass 4") ? '<span class="usj-decision">本行程首选</span>' : ""}
      </div>
      <div class="usj-strategy-meta">
        <span><strong>费用：</strong>${item["追加费用"]}</span>
        <span><strong>预计：</strong>${item["预计成果"]}</span>
      </div>
      <div class="field"><strong>优点</strong><p>${item["优点"]}</p></div>
      <div class="field"><strong>缺点</strong><p>${item["缺点"]}</p></div>
      <div class="field"><strong>适合</strong><p>${item["适合"]}</p></div>
    </article>
  `).join("");

  const expressPass4 = guide.express_pass4.map((item, index) => `
    <article class="usj-pass-card ${index === 0 ? "is-recommended" : ""}">
      <div class="usj-date-top">
        <h4>${item["名称"]}</h4>
        <span class="usj-decision">${item["标签"]}</span>
      </div>
      <ol class="usj-pass-projects">
        ${item["项目"].map(project => `<li>${project}</li>`).join("")}
      </ol>
      <div class="field"><strong>为什么选</strong><p>${item["推荐理由"]}</p></div>
      <div class="field"><strong>情侣玩法</strong><p>${item["情侣建议"]}</p></div>
    </article>
  `).join("");

  const singleRiderCouples = guide.single_rider_couples.map(item => `
    <article class="usj-single-card">
      <div class="usj-date-top">
        <h4>${item["项目"]}</h4>
        <span class="usj-single-status">${item["建议"]}</span>
      </div>
      <p>${item["情侣策略"]}</p>
    </article>
  `).join("");

  const sources = guide.sources.map(item => `
    <a class="usj-source-link" href="${item["网址"]}" target="_blank" rel="noopener noreferrer">${item["名称"]}</a>
  `).join("");

  $("#usjGuide").innerHTML = `
    <article class="usj-hero-card">
      <h3>客流优先：9月27日＋核心Express Pass 4</h3>
      <p>${guide.summary["结论"]}</p>
      <p class="usj-small-note">${guide.summary["预测说明"]}</p>
      <p class="usj-small-note">当天动作：${guide.summary["当天动作"]}</p>
    </article>

    <section class="usj-section">
      <div class="usj-section-head"><h3>27、28还是29日</h3><span>当前预测，天气变化后需重查</span></div>
      <div class="usj-date-grid">${dates}</div>
    </section>

    <section class="usj-section usj-map-section">
      <div class="usj-section-head"><h3>园区大地图＋路线编号</h3><span>手机可横向滑动查看</span></div>
      ${renderUSJRouteMap()}
      <a class="usj-official-map-link" href="https://www.usj.co.jp/web/ja/jp/service-guide/parkmap" target="_blank" rel="noopener noreferrer">打开USJ官方实时地图</a>
    </section>

    <section class="usj-section">
      <div class="usj-section-head"><h3>主要项目</h3><span>点击项目展开介绍与取舍</span></div>
      <div class="usj-attraction-list">${attractions}</div>
    </section>

    <section class="usj-section">
      <div class="usj-section-head"><h3>万圣节限定</h3><span>9月27—29日都有普通万圣惊魂夜</span></div>
      <div class="usj-halloween-list">${halloween}</div>
    </section>

    <section class="usj-section">
      <div class="usj-section-head"><h3>Express Pass 4 怎么选</h3><span>两人各买一张，并一次下单</span></div>
      <div class="usj-pass-grid">${expressPass4}</div>
      <p class="usj-inline-note">套餐名称、项目和指定时段可能调整；购买9月27日票券时，以结算页列出的4个项目及区域入场保证为最终依据。</p>
    </section>

    <section class="usj-section">
      <div class="usj-section-head"><h3>速通4条件下的最优时间轴</h3><span>${guide.route_assumption["适用日期"]}</span></div>
      <div class="usj-route-assumption">
        <strong>${guide.route_assumption["目标套餐"]}</strong>
        <p>建议购票时段：${guide.route_assumption["建议时段"]}</p>
        <p>${guide.route_assumption["固定节点"]}</p>
        <p>${guide.route_assumption["说明"]}</p>
      </div>
      <div class="usj-route-timeline">${routeTimeline}</div>
    </section>

    <section class="usj-section">
      <div class="usj-section-head"><h3>情侣使用单人通道</h3><span>省时间，但必须接受拆开乘坐</span></div>
      <div class="usj-single-grid">${singleRiderCouples}</div>
      <ul class="usj-closure-list usj-single-rules">${guide.single_rider_rules.map(item => `<li>${item}</li>`).join("")}</ul>
    </section>

    <section class="usj-section">
      <div class="usj-section-head"><h3>速通购买方案</h3><span>入园票需要另买</span></div>
      <div class="usj-strategy-grid">${strategies}</div>
    </section>

    <section class="usj-section">
      <div class="usj-section-head"><h3>休止与临场风险</h3><span>出发前一晚再检查</span></div>
      <ul class="usj-closure-list">${guide.closures.map(item => `<li>${item}</li>`).join("")}</ul>
    </section>

    <section class="usj-section">
      <div class="usj-section-head"><h3>官网与数据入口</h3><span>手机可直接打开</span></div>
      <div class="usj-source-list">${sources}</div>
    </section>
  `;
}

function renderSimpleLists() {
  $("#lodgingList").innerHTML = data.lodging.map(lodgingCard).join("");

  $("#planbList").innerHTML = data.alternatives.map(item => simpleCard(
    item["场景"],
    [item["原计划"]],
    [["改法", item["改法"]], ["收益", item["收益"]], ["备注", item["备注"]]]
  )).join("");

  $("#transportList").innerHTML = data.transport.map(item => simpleCard(
    item["事项"],
    [item["建议时间"]],
    [["建议做法", item["建议做法"]], ["风险点", item["风险点"]], ["P人提醒", item["P人提醒"]]]
  )).join("");

  $("#checkList").innerHTML = data.checklist.map(item => simpleCard(
    `${item["类别"]} · ${item["动作"]}`,
    [item["优先级"]],
    [["建议", item["建议"]], ["为什么", item["为什么"]]]
  )).join("");
}

function switchView(view) {
  state.view = view;
  document.querySelectorAll(".tab").forEach(tab => tab.classList.toggle("is-active", tab.dataset.view === view));
  document.querySelectorAll(".content-view").forEach(panel => panel.classList.remove("is-visible"));
  $(`#${view}View`).classList.add("is-visible");
  $("#dayToolbar").style.display = view === "days" ? "block" : "none";
}

function viewFromHash() {
  const requested = window.location.hash.replace("#", "");
  const validViews = ["days", "food", "lodging", "usj", "planb", "prep"];
  return validViews.includes(requested) ? requested : "days";
}

function bindEvents() {
  document.querySelector(".tabs").addEventListener("click", (event) => {
    const button = event.target.closest(".tab");
    if (!button) return;
    switchView(button.dataset.view);
    history.replaceState(null, "", `#${button.dataset.view}`);
  });

  $("#cityChips").addEventListener("click", (event) => {
    const button = event.target.closest(".chip");
    if (!button) return;
    state.city = button.dataset.city;
    renderChips();
    renderDays();
  });

  $("#searchInput").addEventListener("input", (event) => {
    state.query = event.target.value.trim();
    renderDays();
  });

  window.addEventListener("hashchange", () => switchView(viewFromHash()));
}

renderHero();
renderChips();
renderDays();
renderFood();
renderUSJ();
renderSimpleLists();
bindEvents();
const initialView = viewFromHash();
switchView(initialView);
if (initialView !== "days") {
  requestAnimationFrame(() => document.querySelector(".tabs").scrollIntoView());
}
""",
        encoding="utf-8",
    )


def write_standalone() -> None:
    html = (ROOT / "index.html").read_text(encoding="utf-8")
    css = (ROOT / "styles.css").read_text(encoding="utf-8")
    data_js = (ROOT / "data.js").read_text(encoding="utf-8")
    app_js = (ROOT / "app.js").read_text(encoding="utf-8")

    html = html.replace('    <link rel="stylesheet" href="./styles.css" />', f"    <style>\n{css}\n    </style>")
    html = html.replace('    <script src="./data.js"></script>\n    <script src="./app.js"></script>', f"    <script>\n{data_js}\n{app_js}\n    </script>")
    html = html.replace(
        '<link rel="preconnect" href="https://images.unsplash.com" />\n    ',
        "",
    )
    # GitHub Pages 从仓库根目录的 index.html 进入；同时保留原中文文件名供本地直接打开。
    Path("index.html").write_text(html, encoding="utf-8")
    Path("日本12天11晚_手机单页攻略.html").write_text(html, encoding="utf-8")


if __name__ == "__main__":
    write_site(build_data())
    write_standalone()
    print((ROOT / "index.html").resolve())
    print(Path("index.html").resolve())
    print(Path("日本12天11晚_手机单页攻略.html").resolve())
